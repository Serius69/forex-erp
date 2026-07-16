"""
Sincroniza las transacciones del Google Sheet operativo (KapitalyaRegistro2026)
hacia la BD — el ETL que faltaba: antes la carga era 100% manual (última:
2026-07-10) y la infraestructura de data_migration estaba muerta por falta de
credenciales de service account.

Cómo funciona (SIN credenciales — la hoja permite export por link):
  1. Descarga el libro completo como XLSX (las fechas llegan TIPADAS como
     datetime — evita el clásico problema D/M/Y vs M/D/Y del export CSV).
  2. Lee la pestaña 'Transacciones' (seq, Fecha, Tipo, Divisa, Cantidad,
     TC, Total Bs, Medio Pago, Responsable, Notas).
  3. Idempotencia por FILA: cada tx insertada guarda 'GS2026#<seq>' en
     payment_reference; una seq ya presente jamás se re-inserta. Las filas
     históricas pre-marcador (cargas manuales ≤ 2026-07-10) se respetan vía
     fecha de corte auto-detectada.
  4. Inserta con bulk_create (bypassa las 7 señales post_save: sin RTE/alertas/
     caja retroactivos), categoría INTERNA + visible_asfi=False (convención de
     las cargas previas) y BACKDATEA created_at al timestamp real de la hoja.

Uso:
    python manage.py sync_sheet_transactions              # sincroniza
    python manage.py sync_sheet_transactions --dry-run    # solo reporta
    python manage.py sync_sheet_transactions --since 2026-07-10
"""
import hashlib
import io
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand
from django.utils import timezone

# ID del libro 2026 (export por link, sin credenciales). Sobre-escribible por
# settings.FOREX_SHEET_2026_ID o --sheet-id.
DEFAULT_SHEET_ID = '1ZAL08c671-3jDAATgOn7MpqngwswKBh4yXIssBTP4xE'
TAB_NAME = 'Transacciones'
MARK_PREFIX = 'GS2026#'
NOTES_MARK = 'CARGA_GS_SYNC'

# Cargas manuales previas (para detectar la fecha de corte de la transición)
LEGACY_MARKS = ('CARGA_HISTORICA_CSV', 'CARGA_GAP_2026_SHEET')

PAYMENT_MAP = {
    'efectivo': 'CASH',
    'transferencia': 'TRANSFER',
    'qr': 'QR',
    'tarjeta': 'CARD',
}

# Alias de divisa usados en la hoja → código de Currency del sistema
# (las variantes de caja son productos operativos con su propia tasa)
DIVISA_ALIAS = {
    'USD SUELTOS':  'USD_CASH_LOOSE',
    'USD 1 Y 2':    'USD_SMALL_BILLS',
    'PEN MON':      'PEN_COINS',
    'PEN MONEDAS':  'PEN_COINS',
}


def _num(val):
    """float/int/str('4.000,00') → Decimal, o None."""
    if val is None or val == '':
        return None
    if isinstance(val, (int, float)):
        return Decimal(str(val))
    s = str(val).strip()
    if ',' in s and '.' in s:          # 4.000,00 → 4000.00
        s = s.replace('.', '').replace(',', '.')
    elif ',' in s:
        s = s.replace(',', '.')
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


class Command(BaseCommand):
    help = 'Sincroniza transacciones del Google Sheet 2026 → BD (idempotente).'

    def add_arguments(self, parser):
        parser.add_argument('--sheet-id', default=None)
        parser.add_argument('--dry-run', action='store_true')
        parser.add_argument('--since', default=None,
                            help='Solo filas con fecha > YYYY-MM-DD (default: auto).')

    # ── Descarga y parseo ────────────────────────────────────────────────────
    def _fetch_workbook(self, sheet_id):
        import openpyxl
        import requests

        url = f'https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx'
        resp = requests.get(url, timeout=60)
        resp.raise_for_status()
        ct = resp.headers.get('content-type', '')
        if 'spreadsheetml' not in ct:
            raise RuntimeError(
                f'Export inesperado ({ct[:60]}) — ¿la hoja dejó de estar '
                'compartida por link?')
        return openpyxl.load_workbook(io.BytesIO(resp.content),
                                      read_only=True, data_only=True)

    def _fetch_rows(self, wb):
        if TAB_NAME not in wb.sheetnames:
            raise RuntimeError(f'Pestaña {TAB_NAME!r} no encontrada')
        ws = wb[TAB_NAME]

        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            seq, fecha, _mes, tipo, divisa = row[0], row[1], row[2], row[3], row[4]
            if fecha is None or tipo is None:
                continue
            if not isinstance(fecha, datetime):
                # Con XLSX no debería pasar; si pasa, registrar y saltar (no adivinar D/M/Y)
                self.stderr.write(f'  fila con fecha no tipada, saltada: {fecha!r}')
                continue
            rows.append({
                'seq':       int(float(seq)) if seq is not None else None,
                'fecha':     fecha,
                'tipo':      str(tipo).strip().lower(),
                'divisa':    str(divisa or '').strip().upper(),
                'cantidad':  _num(row[5]),
                'tc':        _num(row[6]),
                'total_bs':  _num(row[7]),
                'medio':     str(row[8] or 'Efectivo').strip().lower(),
                'resp':      str(row[9] or '').strip(),
                'notas':     str(row[10] or '').strip(),
            })
        return rows

    # ── Sincronización ───────────────────────────────────────────────────────
    def handle(self, *args, **opts):
        from django.conf import settings as dj_settings
        from django.db.models import Q

        from rates.models import Currency
        from transactions.models import Transaction

        sheet_id = (opts['sheet_id']
                    or getattr(dj_settings, 'FOREX_SHEET_2026_ID', None)
                    or DEFAULT_SHEET_ID)
        dry = opts['dry_run']

        wb = self._fetch_workbook(sheet_id)
        rows = self._fetch_rows(wb)
        self.stdout.write(f'Hoja: {len(rows)} filas de transacciones')

        # Flujos hermanos del mismo libro (idempotentes, no bloquean el principal)
        try:
            self._sync_capital(wb, dry)
        except Exception as exc:
            self.stderr.write(f'capital sync falló: {exc}')
        try:
            self._sync_gastos(wb, dry)
        except Exception as exc:
            self.stderr.write(f'gastos sync falló: {exc}')
        try:
            self._sync_competencia_rates(wb, dry)
        except Exception as exc:
            self.stderr.write(f'competencia rates sync falló: {exc}')

        # Fecha de corte: lo manual previo llegó hasta cierto día completo.
        if opts['since']:
            cutoff = date.fromisoformat(opts['since'])
        else:
            last_legacy = (Transaction.objects
                           .filter(notes__in=LEGACY_MARKS)
                           .order_by('-created_at').first())
            cutoff = last_legacy.created_at.date() if last_legacy else date(2000, 1, 1)
        self.stdout.write(f'Corte legacy: filas con fecha > {cutoff}')

        # seqs ya sincronizadas (marcador estable por fila)
        seen = set(
            Transaction.objects
            .filter(payment_reference__startswith=MARK_PREFIX)
            .values_list('payment_reference', flat=True)
        )

        currencies = {c.code: c for c in Currency.objects.all()}
        bob = currencies.get('BOB')
        if bob is None:
            self.stderr.write(self.style.ERROR('Falta BOB — corre seed_currencies'))
            return

        cashier = self._default_cashier()
        branch = cashier.branch if cashier and cashier.branch_id else None
        if branch is None:
            from users.models import Branch
            branch = Branch.objects.filter(is_active=True).order_by('id').first()

        # Mapa Responsable (hoja) → usuario real. La columna "Responsable" suele
        # traer el NOMBRE VISIBLE (no el username), así que indexamos por username,
        # first_name, last_name y nombre completo. El usuario unificado SSO 'sergio'
        # (kapitalyabolivia@gmail.com) queda como fallback.
        from django.contrib.auth import get_user_model
        User = get_user_model()
        cashier_map: dict[str, object] = {}
        for u in User.objects.filter(is_active=True):
            keys = {
                (u.username or '').strip().lower(),
                (u.first_name or '').strip().lower(),
                (u.last_name or '').strip().lower(),
                f"{(u.first_name or '').strip()} {(u.last_name or '').strip()}".strip().lower(),
                (u.get_full_name() or '').strip().lower(),
            }
            for k in keys:
                if k and k not in cashier_map:  # el username tiene prioridad (se agrega primero)
                    cashier_map[k] = u

        def _cashier_for(resp: str):
            key = (resp or '').strip().lower()
            if key in cashier_map:
                return cashier_map[key]
            # coincidencia por primer token del nombre visible (p.ej. "Sergio T.")
            first = key.split()[0] if key.split() else ''
            return cashier_map.get(first, cashier)

        # Marcador estable por fila para deduplicar entre corridas. Con seq usa
        # el número; SIN seq (filas que no traen correlativo) deriva un hash del
        # contenido para NO re-insertarlas en cada sync (antes duplicaban).
        def _row_mark(r) -> str:
            if r['seq'] is not None:
                return f"{MARK_PREFIX}{r['seq']}"
            basis = (f"{r['fecha'].isoformat()}|{r['divisa']}|{r['cantidad']}"
                     f"|{r['tc']}|{r['resp']}")
            return f"{MARK_PREFIX}H{hashlib.sha1(basis.encode()).hexdigest()[:12]}"

        pendientes, saltadas = [], 0
        for r in rows:
            if r['fecha'].date() <= cutoff:
                continue
            mark = _row_mark(r)
            if mark in seen:
                continue
            seen.add(mark)  # evita duplicar filas idénticas dentro del MISMO batch

            code = DIVISA_ALIAS.get(r['divisa'], r['divisa'])
            cur = currencies.get(code)
            if cur is None or r['cantidad'] is None or r['tc'] is None:
                saltadas += 1
                self.stderr.write(f"  saltada seq={r['seq']}: divisa/números inválidos {r['divisa']}")
                continue
            total = r['total_bs'] if r['total_bs'] is not None else (r['cantidad'] * r['tc'])
            tipo = 'SELL' if r['tipo'].startswith('venta') else 'BUY'
            pendientes.append((r, mark, cur, tipo, total))

        self.stdout.write(f'Nuevas a insertar: {len(pendientes)} · saltadas: {saltadas}')
        if dry or not pendientes:
            if dry:
                for r, mark, cur, tipo, total in pendientes[:15]:
                    self.stdout.write(
                        f"  {r['fecha']} {tipo} {r['cantidad']} {cur.code} "
                        f"@{r['tc']} = Bs {total} [{mark}]")
                self.stdout.write(self.style.WARNING('--dry-run: sin cambios.'))
            else:
                self.stdout.write(self.style.SUCCESS('Nada nuevo que sincronizar.'))
            return

        # Numeración gap por día: 01YYYYMMDDG#### continuando la seq existente
        def _next_number(d, counters={}):
            key = d.strftime('%Y%m%d')
            if key not in counters:
                prefix = f'01{key}G'
                last = (Transaction.objects
                        .filter(transaction_number__startswith=prefix)
                        .order_by('-transaction_number').first())
                counters[key] = int(last.transaction_number[-4:]) if last else 0
            counters[key] += 1
            return f'01{key}G{counters[key]:04d}'

        objs, fechas = [], []
        for r, mark, cur, tipo, total in sorted(pendientes, key=lambda p: p[0]['fecha']):
            fecha_aw = (timezone.make_aware(r['fecha'])
                        if timezone.is_naive(r['fecha']) else r['fecha'])
            objs.append(Transaction(
                transaction_number   = _next_number(fecha_aw),
                transaction_type     = tipo,
                transaction_category = 'INTERNA',
                status               = 'COMPLETED',
                currency_from        = cur,
                currency_to          = bob,
                amount_from          = int(round(float(r['cantidad']))),
                amount_to            = int(round(float(total))),
                exchange_rate        = r['tc'].quantize(Decimal('0.0001')),
                payment_method       = PAYMENT_MAP.get(r['medio'], 'CASH'),
                payment_reference    = mark or '',
                notes                = (f"{NOTES_MARK} resp={r['resp']}"
                                        + (f" | {r['notas']}" if r['notas'] else '')),
                visible_asfi         = False,
                cashier              = _cashier_for(r['resp']),
                branch               = branch,
                completed_at         = fecha_aw,
            ))
            fechas.append(fecha_aw)

        # bulk_create bypassa señales post_save (RTE/caja/alertas retroactivos)
        created = Transaction.objects.bulk_create(objs, batch_size=200)
        # Backdatear created_at (auto_now_add lo pisó al insertar)
        for obj, f in zip(created, fechas):
            Transaction.objects.filter(pk=obj.pk).update(created_at=f, updated_at=f)

        self.stdout.write(self.style.SUCCESS(
            f'Sincronizadas {len(created)} transacciones '
            f'({fechas[0].date()} → {fechas[-1].date()}).'))

    @staticmethod
    def _default_cashier():
        from django.contrib.auth import get_user_model
        User = get_user_model()
        return (User.objects.filter(username='sergio', is_active=True).first()
                or User.objects.filter(is_active=True, role='ADMIN').first()
                or User.objects.filter(is_active=True).first())

    # ── Balances de capital (pestaña 'Composicion Capital', bloque timeline) ──
    def _sync_capital(self, wb, dry):
        from capital.models import CapitalSnapshot
        from users.models import Branch

        if 'Composicion Capital' not in wb.sheetnames:
            return
        ws = wb['Composicion Capital']
        branch = Branch.objects.filter(is_active=True).order_by('id').first()
        user = self._default_cashier()   # generado_por es NOT NULL

        nuevos = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            # bloque timeline: Nro(4), Fecha(5), Mes(6), Balance(7)
            nro, fecha, balance = row[4], row[5], row[7]
            if not isinstance(fecha, datetime) or balance is None:
                continue
            total = _num(balance)
            if total is None or total <= 0:
                continue
            f = fecha.date()
            # dedupe por (fecha, total redondeado) — mismo criterio que la
            # carga manual previa (varias filas por fecha son legítimas)
            if CapitalSnapshot.objects.filter(
                    fecha=f, total_bob=total.quantize(Decimal('0.01'))).exists():
                continue
            nuevos += 1
            if not dry:
                CapitalSnapshot.objects.create(
                    fecha=f, branch=branch, generado_por=user,
                    total_bob=total.quantize(Decimal('0.01')),
                    notas=(f'{NOTES_MARK} Nro={int(float(nro)) if nro else "?"} — '
                           'balance total de "Composicion Capital" (solo total)'),
                )
        self.stdout.write(f'Capital: {nuevos} snapshot(s) nuevo(s)'
                          + (' [dry-run]' if dry and nuevos else ''))

    # ── Gastos (pestaña 'CATEGORIZACIÓN DE GASTOS') ───────────────────────────
    def _sync_gastos(self, wb, dry):
        from capital.models import Gasto
        from users.models import Branch

        tab = 'CATEGORIZACIÓN DE GASTOS'
        if tab not in wb.sheetnames:
            return
        ws = wb[tab]
        branch = Branch.objects.filter(is_active=True).order_by('id').first()

        # marcador estable por fila de la hoja (para syncs futuros); las cargas
        # legacy (sin marcador) se cubren con el corte por fecha
        ultimo = Gasto.objects.order_by('-fecha').first()
        cutoff = ultimo.fecha if ultimo else date(2000, 1, 1)
        marcados = set(Gasto.objects.filter(notas__startswith='GS_GASTO#')
                       .values_list('notas', flat=True))

        nuevos = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            gid, fecha, _mes, cat, monto, medio, coment, resp = (row + (None,) * 8)[:8]
            if not isinstance(fecha, datetime) or monto is None:
                continue
            mark = f'GS_GASTO#{int(float(gid))}' if gid is not None else None
            if mark and any(m.startswith(mark + ' ') or m == mark for m in marcados):
                continue
            if fecha.date() <= cutoff and not mark:
                continue
            if fecha.date() <= cutoff:
                # legacy sin marcador ya cargado — solo insertar si es posterior
                continue
            m = _num(monto)
            if m is None or m <= 0:
                continue
            nuevos += 1
            if not dry:
                Gasto.objects.create(
                    fecha=fecha.date(),
                    categoria=str(cat or 'Otros')[:100],
                    descripcion=str(coment or '')[:255],
                    monto_bob=m.quantize(Decimal('0.01')),
                    medio_pago=str(medio or '')[:50],
                    notas=f'{mark or "GS_GASTO#?"} {NOTES_MARK} resp={resp or ""}',
                    branch=branch,
                )
        self.stdout.write(f'Gastos: {nuevos} nuevo(s)'
                          + (' [dry-run]' if dry and nuevos else ''))

    # ── Tasas físicas de COMPETENCIA (revive la serie congelada en marzo) ─────
    # Fuentes del mismo libro:
    #   · 'TCHistoricoCompetencia': backfill 2026-03-31 → hoy (observaciones
    #     manuales del mercado físico — cubre el hueco del CSV histórico)
    #   · 'TC Mercado HOY': la foto vigente (se re-observa en cada sync → la
    #     serie competencia queda VIVA, ya no depende de cargas manuales)
    def _sync_competencia_rates(self, wb, dry):
        from django.utils import timezone as tz

        from rates.models import Currency, ExchangeRate

        COMPETENCIA = 'paralelo_fisico_competencia'
        SRC = 'GS_TC_COMPETENCIA'
        currencies = {c.code: c for c in Currency.objects.all()}
        bob = currencies.get('BOB')
        if bob is None:
            return

        def _upsert(code_raw, compra, venta, fecha_dt):
            code = DIVISA_ALIAS.get((code_raw or '').strip().upper(),
                                    (code_raw or '').strip().upper())
            cur = currencies.get(code)
            buy, sell = _num(compra), _num(venta)
            if cur is None or not buy or not sell or buy <= 0 or sell <= 0:
                return 0
            if buy > sell:
                buy, sell = sell, buy
            valid_from = (tz.make_aware(fecha_dt)
                          if tz.is_naive(fecha_dt) else fecha_dt)
            q4 = Decimal('0.0001')
            if dry:
                exists = ExchangeRate.objects.filter(
                    currency_from=cur, currency_to=bob, valid_from=valid_from,
                    market_type=COMPETENCIA, rate_source=None).exists()
                return 0 if exists else 1
            _, created = ExchangeRate.objects.update_or_create(
                currency_from=cur, currency_to=bob, valid_from=valid_from,
                market_type=COMPETENCIA, rate_source=None,
                defaults={
                    'official_rate': ((buy + sell) / 2).quantize(q4),
                    'buy_rate':  buy.quantize(q4),
                    'sell_rate': sell.quantize(q4),
                    'avg_rate':  ((buy + sell) / 2).quantize(q4),
                    'source': SRC,
                    'source_method': 'MANUAL',
                    'confidence': Decimal('0.85'),
                },
            )
            return int(created)

        nuevos = 0

        # 1) Backfill histórico (Fecha, Divisa, Compra, Venta, Promedio, …)
        if 'TCHistoricoCompetencia' in wb.sheetnames:
            for row in wb['TCHistoricoCompetencia'].iter_rows(min_row=2, values_only=True):
                fecha, divisa, compra, venta = row[0], row[1], row[2], row[3]
                if not isinstance(fecha, datetime):
                    continue
                nuevos += _upsert(divisa, compra, venta, fecha)

        # 2) Foto vigente (Divisa, Compra, Venta, Promedio, Última Actualización)
        if 'TC Mercado HOY' in wb.sheetnames:
            for row in wb['TC Mercado HOY'].iter_rows(min_row=2, values_only=True):
                divisa, compra, venta = row[0], row[1], row[2]
                ts = row[4] if len(row) > 4 else None
                if not divisa or not isinstance(ts, datetime):
                    continue
                # anclar al DÍA de la última actualización (una obs/día)
                dia = datetime(ts.year, ts.month, ts.day)
                nuevos += _upsert(divisa, compra, venta, dia)

        if nuevos and not dry:
            # dejar UNA vigente por divisa (intervalos contiguos)
            from rates.rate_expiry import expire_stale_active_rates
            expire_stale_active_rates(market_type=COMPETENCIA)
        self.stdout.write(f'Competencia: {nuevos} tasa(s) nueva(s)'
                          + (' [dry-run]' if dry and nuevos else ''))
