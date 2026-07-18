# backend/reports/generators.py
import io
import os
from datetime import datetime, time, timedelta
from decimal import Decimal
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.lineplots import LinePlot
from reportlab.graphics.charts.barcharts import VerticalBarChart
from django.core.files.base import ContentFile
from django.db.models import Sum, Count, Avg, Q, F
from django.db.models.functions import ExtractHour
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference


def _day_bounds(day):
    """Un día (TZ local) → límites datetime sargables (lo, hi) con lo <= created_at < hi.
    Evita el DATE(created_at AT TIME ZONE ...) no-sargable de created_at__date=."""
    tz = timezone.get_current_timezone()
    lo = timezone.make_aware(datetime.combine(day, time.min), tz)
    hi = timezone.make_aware(datetime.combine(day + timedelta(days=1), time.min), tz)
    return lo, hi


class ReportGenerator:
    """Generador principal de reportes"""
    
    def __init__(self, start_date, end_date, branch=None, user=None):
        self.start_date = start_date
        self.end_date = end_date
        self.branch = branch
        self.user = user
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()
    
    def _setup_custom_styles(self):
        """Configura estilos personalizados"""
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1976d2'),
            spaceAfter=30,
            alignment=TA_CENTER
        ))
        
        self.styles.add(ParagraphStyle(
            name='SectionTitle',
            parent=self.styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#333333'),
            spaceAfter=12
        ))
    
    def generate_daily_report(self):
        """Genera reporte diario completo"""
        from transactions.models import Transaction, Customer
        from inventory.models import CurrencyInventory, InventoryMovement
        from rates.models import ExchangeRate
        
        # Recopilar datos
        _lo, _hi = _day_bounds(self.start_date)
        transactions = Transaction.objects.filter(
            created_at__gte=_lo, created_at__lt=_hi
        )
        
        if self.branch:
            transactions = transactions.filter(branch=self.branch)
        
        # Estadísticas generales
        stats = {
            'total_transactions': transactions.count(),
            'total_customers': transactions.values('customer').distinct().count(),
            'total_volume_bob': transactions.aggregate(
                total=Sum('amount_to')
            )['total'] or 0,
            # order_by() explícito: el ordering default (-created_at) contamina
            # el GROUP BY y produce un grupo por fila
            'by_type': transactions.values('transaction_type').order_by('transaction_type').annotate(
                count=Count('id'),
                volume=Sum('amount_to')
            ),
            'by_currency': transactions.values('currency_from__code').order_by('currency_from__code').annotate(
                count=Count('id'),
                volume=Sum('amount_from'),
                volume_bob=Sum('amount_to'),
                buy_count=Count('id', filter=Q(transaction_type='BUY')),
                sell_count=Count('id', filter=Q(transaction_type='SELL'))
            ),
            'by_payment': transactions.values('payment_method').order_by('payment_method').annotate(
                count=Count('id'),
                volume=Sum('amount_to')
            ),
            'by_hour': self._get_hourly_distribution(transactions),
            'top_customers': self._get_top_customers(transactions),
            'profit_analysis': self._calculate_daily_profit(transactions),
            'inventory_status': self._get_inventory_status()
        }
        
        # Generar PDF
        pdf_buffer = self._generate_daily_pdf(stats, transactions)
        
        # Generar Excel
        excel_buffer = self._generate_daily_excel(stats, transactions)
        
        # Crear registro de reporte
        from .models import Report
        
        report = Report.objects.create(
            report_type='DAILY',
            title=f'Reporte Diario - {self.start_date}',
            description=f'Reporte diario de operaciones para {self.branch.name if self.branch else "todas las sucursales"}',
            start_date=self.start_date,
            end_date=self.start_date,
            pdf_file=ContentFile(pdf_buffer.getvalue(), name=f'daily_{self.start_date}.pdf'),
            excel_file=ContentFile(excel_buffer.getvalue(), name=f'daily_{self.start_date}.xlsx'),
            summary_data=stats,
            generated_by=self.user
        )
        
        return report
    
    def _generate_daily_pdf(self, stats, transactions):
        """Genera PDF del reporte diario"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        
        # Título
        story.append(Paragraph("REPORTE DIARIO DE OPERACIONES", self.styles['CustomTitle']))
        story.append(Paragraph(f"Fecha: {self.start_date}", self.styles['Normal']))
        if self.branch:
            story.append(Paragraph(f"Sucursal: {self.branch.name}", self.styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Resumen ejecutivo
        story.append(Paragraph("RESUMEN EJECUTIVO", self.styles['SectionTitle']))
        
        summary_data = [
            ['Indicador', 'Valor'],
            ['Total Transacciones', f"{stats['total_transactions']:,}"],
            ['Clientes Atendidos', f"{stats['total_customers']:,}"],
            ['Volumen Total (BOB)', f"Bs. {stats['total_volume_bob']:,.2f}"],
            ['Utilidad Estimada', f"Bs. {stats['profit_analysis']['total_profit']:,.2f}"],
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1976d2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Desglose por tipo de transacción
        story.append(Paragraph("DESGLOSE POR TIPO DE TRANSACCIÓN", self.styles['SectionTitle']))
        
        type_data = [['Tipo', 'Cantidad', 'Volumen (BOB)']]
        for item in stats['by_type']:
            type_data.append([
                'Compra' if item['transaction_type'] == 'BUY' else 'Venta',
                f"{item['count']:,}",
                f"Bs. {item['volume']:,.2f}"
            ])
        
        type_table = Table(type_data, colWidths=[2*inch, 1.5*inch, 2*inch])
        type_table.setStyle(self._get_table_style())
        story.append(type_table)
        story.append(Spacer(1, 20))
        
        # Desglose por divisa
        story.append(Paragraph("OPERACIONES POR DIVISA", self.styles['SectionTitle']))
        
        currency_data = [['Divisa', 'Transacciones', 'Volumen', 'Volumen (BOB)']]
        for item in stats['by_currency']:
            currency_data.append([
                item['currency_from__code'],
                f"{item['count']:,}",
                f"{item['volume']:,.2f}",
                f"Bs. {item['volume_bob']:,.2f}"
            ])
        
        currency_table = Table(currency_data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
        currency_table.setStyle(self._get_table_style())
        story.append(currency_table)
        
        # Gráfico de distribución horaria
        story.append(PageBreak())
        story.append(Paragraph("DISTRIBUCIÓN HORARIA", self.styles['SectionTitle']))
        
        # Crear gráfico
        hourly_chart = self._create_hourly_chart(stats['by_hour'])
        story.append(hourly_chart)
        
        # Top clientes
        story.append(Spacer(1, 20))
        story.append(Paragraph("TOP 10 CLIENTES", self.styles['SectionTitle']))
        
        customer_data = [['Cliente', 'Documento', 'Transacciones', 'Volumen (BOB)']]
        for customer in stats['top_customers'][:10]:
            customer_data.append([
                customer['customer__full_name'][:30],
                customer['customer__document_number'],
                f"{customer['count']:,}",
                f"Bs. {customer['total_volume']:,.2f}"
            ])
        
        customer_table = Table(customer_data)
        customer_table.setStyle(self._get_table_style())
        story.append(customer_table)
        
        # Estado del inventario
        story.append(PageBreak())
        story.append(Paragraph("ESTADO DEL INVENTARIO", self.styles['SectionTitle']))
        
        inventory_data = [['Divisa', 'Saldo Inicial', 'Entradas', 'Salidas', 'Saldo Final', 'Estado']]
        for inv in stats['inventory_status']:
            status = 'Bajo' if inv['needs_replenishment'] else 'Normal'
            if inv['is_overstocked']:
                status = 'Exceso'
            
            inventory_data.append([
                inv['currency'],
                f"{inv['initial_balance']:,.2f}",
                f"{inv['total_in']:,.2f}",
                f"{inv['total_out']:,.2f}",
                f"{inv['final_balance']:,.2f}",
                status
            ])
        
        inventory_table = Table(inventory_data)
        inventory_table.setStyle(self._get_table_style())
        story.append(inventory_table)
        
        # Construir PDF
        doc.build(story)
        buffer.seek(0)
        
        return buffer
    
    def _generate_daily_excel(self, stats, transactions):
        """Genera Excel del reporte diario"""
        wb = Workbook()
        
        # Hoja de resumen
        ws_summary = wb.active
        ws_summary.title = "Resumen"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws_summary['A1'] = f"REPORTE DIARIO - {self.start_date}"
        ws_summary['A1'].font = Font(size=16, bold=True)
        ws_summary.merge_cells('A1:E1')
        
        # Resumen
        row = 3
        ws_summary[f'A{row}'] = "RESUMEN EJECUTIVO"
        ws_summary[f'A{row}'].font = Font(size=14, bold=True)
        
        row += 2
        summary_items = [
            ('Total Transacciones', stats['total_transactions']),
            ('Clientes Atendidos', stats['total_customers']),
            ('Volumen Total (BOB)', stats['total_volume_bob']),
            ('Utilidad Estimada', stats['profit_analysis']['total_profit']),
        ]
        
        for label, value in summary_items:
            ws_summary[f'A{row}'] = label
            ws_summary[f'B{row}'] = value
            ws_summary[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Hoja de transacciones
        ws_trans = wb.create_sheet("Transacciones")
        
        # Headers
        trans_headers = [
            'Número', 'Fecha/Hora', 'Tipo', 'Cliente', 'Documento',
            'Divisa', 'Monto', 'Tasa', 'Total BOB', 'Método Pago',
            'Cajero', 'Estado'
        ]
        
        for col, header in enumerate(trans_headers, 1):
            cell = ws_trans.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Datos de transacciones
        row = 2
        for trans in transactions.select_related('customer', 'currency_from', 'cashier'):
            ws_trans.cell(row=row, column=1, value=trans.transaction_number)
            ws_trans.cell(row=row, column=2, value=trans.created_at.strftime('%Y-%m-%d %H:%M'))
            ws_trans.cell(row=row, column=3, value=trans.get_transaction_type_display())
            ws_trans.cell(row=row, column=4, value=trans.customer.full_name)
            ws_trans.cell(row=row, column=5, value=trans.customer.document_number)
            ws_trans.cell(row=row, column=6, value=trans.currency_from.code)
            ws_trans.cell(row=row, column=7, value=float(trans.amount_from))
            ws_trans.cell(row=row, column=8, value=float(trans.exchange_rate))
            ws_trans.cell(row=row, column=9, value=float(trans.amount_to))
            ws_trans.cell(row=row, column=10, value=trans.get_payment_method_display())
            ws_trans.cell(row=row, column=11, value=trans.cashier.get_full_name())
            ws_trans.cell(row=row, column=12, value=trans.get_status_display())
            row += 1
        
        # Ajustar anchos de columna
        for column in ws_trans.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_trans.column_dimensions[column_letter].width = adjusted_width
        
        # Hoja de análisis por divisa
        ws_currency = wb.create_sheet("Análisis por Divisa")
        
        # Headers
        currency_headers = ['Divisa', 'Compras', 'Ventas', 'Total Trans.', 'Volumen', 'Volumen BOB']
        for col, header in enumerate(currency_headers, 1):
            cell = ws_currency.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Datos
        row = 2
        for item in stats['by_currency']:
            # Compras y ventas ya vienen anotadas en el aggregate de by_currency
            ws_currency.cell(row=row, column=1, value=item['currency_from__code'])
            ws_currency.cell(row=row, column=2, value=item['buy_count'])
            ws_currency.cell(row=row, column=3, value=item['sell_count'])
            ws_currency.cell(row=row, column=4, value=item['count'])
            ws_currency.cell(row=row, column=5, value=float(item['volume']))
            ws_currency.cell(row=row, column=6, value=float(item['volume_bob']))
            row += 1
        
        # Agregar gráfico
        chart = BarChart()
        chart.title = "Transacciones por Divisa"
        chart.style = 13
        chart.x_axis.title = 'Divisa'
        chart.y_axis.title = 'Cantidad'
        
        data = Reference(ws_currency, min_col=2, min_row=1, max_row=row-1, max_col=3)
        cats = Reference(ws_currency, min_col=1, min_row=2, max_row=row-1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws_currency.add_chart(chart, "H2")
        
        # Guardar
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    def _get_hourly_distribution(self, transactions):
        """Obtiene distribución horaria de transacciones"""
        # Una sola query agrupada por hora (ExtractHour respeta TZ La_Paz igual
        # que el lookup __hour); el order_by explícito evita que el ordering
        # default (-created_at) contamine el GROUP BY.
        grouped = (
            transactions
            .annotate(hour=ExtractHour('created_at'))
            .values('hour')
            .order_by('hour')
            .annotate(count=Count('id'), volume=Sum('amount_to'))
        )
        by_hour = {row['hour']: row for row in grouped}

        hourly_data = []
        for hour in range(24):
            row = by_hour.get(hour)
            hourly_data.append({
                'hour': hour,
                'count': row['count'] if row else 0,
                'volume': (row['volume'] if row and row['volume'] is not None else 0),
            })

        return hourly_data
    
    def _get_top_customers(self, transactions):
        """Obtiene los principales clientes"""
        return transactions.values(
            'customer__full_name',
            'customer__document_number'
        ).annotate(
            count=Count('id'),
            total_volume=Sum('amount_to')
        ).order_by('-total_volume')
    
    def _calculate_daily_profit(self, transactions):
        """Calcula la utilidad diaria estimada"""
        total_profit = Decimal('0')
        by_currency = {}
        
        for trans in transactions:
            # Estimar margen de ganancia
            if trans.transaction_type == 'BUY':
                # Casa compra divisas (margen aproximado 0.3%)
                profit = trans.amount_to * Decimal('0.003')
            else:
                # Casa vende divisas (margen aproximado 0.3%)
                profit = trans.amount_to * Decimal('0.003')
            
            total_profit += profit
            
            currency = trans.currency_from.code
            if currency not in by_currency:
                by_currency[currency] = Decimal('0')
            by_currency[currency] += profit
        
        return {
            'total_profit': float(total_profit),
            'by_currency': {k: float(v) for k, v in by_currency.items()}
        }
    
    def _get_inventory_status(self):
        """Obtiene estado del inventario"""
        from inventory.models import CurrencyInventory, InventoryMovement
        
        inventories = CurrencyInventory.objects.all()
        if self.branch:
            inventories = inventories.filter(branch=self.branch)
        
        status_data = []
        
        _lo, _hi = _day_bounds(self.start_date)
        for inv in inventories:
            # Movimientos del día (límites sargables; el N+1 por inventario queda
            # como follow-up: podría precomputarse con values('inventory').annotate(...))
            movements = InventoryMovement.objects.filter(
                inventory=inv,
                created_at__gte=_lo, created_at__lt=_hi
            )
            
            # Calcular saldo inicial (primer movimiento del día)
            first_movement = movements.first()
            if first_movement:
                initial_balance = first_movement.balance_before
            else:
                initial_balance = inv.total_balance
            
            # Sumar entradas y salidas
            total_in = movements.filter(
                movement_type__in=['IN', 'TRANSFER_IN']
            ).aggregate(
                total=Sum('amount')
            )['total'] or Decimal('0')
            
            total_out = movements.filter(
                movement_type__in=['OUT', 'TRANSFER_OUT']
            ).aggregate(
                total=Sum('amount')
            )['total'] or Decimal('0')
            
            status_data.append({
                'currency': inv.currency.code,
                'branch': inv.branch.name,
                'initial_balance': float(initial_balance),
                'total_in': float(total_in),
                'total_out': float(total_out),
                'final_balance': float(inv.total_balance),
                'needs_replenishment': inv.needs_replenishment,
                'is_overstocked': inv.is_overstocked
            })
        
        return status_data
    
    def _create_hourly_chart(self, hourly_data):
        """Crea gráfico de distribución horaria"""
        drawing = Drawing(400, 200)
        
        bc = VerticalBarChart()
        bc.x = 50
        bc.y = 50
        bc.height = 125
        bc.width = 300
        
        # Datos
        data = [[item['count'] for item in hourly_data]]
        bc.data = data
        
        # Categorías (horas)
        bc.categoryAxis.categoryNames = [f"{i}h" for i in range(24)]
        bc.categoryAxis.labels.boxAnchor = 'ne'
        bc.categoryAxis.labels.dx = 8
        bc.categoryAxis.labels.dy = -2
        bc.categoryAxis.labels.angle = 30
        
        bc.valueAxis.valueMin = 0
        bc.valueAxis.valueMax = max([item['count'] for item in hourly_data]) + 5
        bc.valueAxis.valueStep = 5
        
        drawing.add(bc)
        
        return drawing
    
    def _get_table_style(self):
        """Estilo estándar para tablas"""
        return TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1976d2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ])
    
    def generate_regulatory_report(self):
        """Genera reporte para autoridades regulatorias"""
        from transactions.models import Transaction
        
        # Transacciones de alto valor (>$5000 o equivalente)
        high_value_threshold = Decimal('5000')
        bob_threshold = high_value_threshold * Decimal('9.50')  # Tasa paralela aproximada
        
        high_value_transactions = Transaction.objects.filter(
            Q(created_at__range=[self.start_date, self.end_date]) &
            (
                Q(currency_from__code='USD', amount_from__gte=high_value_threshold) |
                Q(amount_to__gte=bob_threshold)
            )
        ).select_related('customer', 'currency_from', 'cashier', 'branch')
        
        # Clientes frecuentes (más de 5 transacciones en el período)
        from django.db.models import Count
        
        frequent_customers = Transaction.objects.filter(
            created_at__range=[self.start_date, self.end_date]
        ).values(
            'customer__id',
            'customer__full_name',
            'customer__document_type',
            'customer__document_number',
            'customer__nationality',
            'customer__is_pep'
        ).annotate(
            transaction_count=Count('id'),
            total_volume_bob=Sum('amount_to')
        ).filter(
            transaction_count__gte=5
        ).order_by('-total_volume_bob')
        
        # Generar PDF del reporte
        buffer = self._generate_regulatory_pdf(
            high_value_transactions,
            frequent_customers
        )
        
        # Crear registro
        from .models import Report
        
        report = Report.objects.create(
            report_type='REGULATORY',
            title=f'Reporte Regulatorio - {self.start_date} a {self.end_date}',
            description='Reporte de cumplimiento para autoridades regulatorias',
            start_date=self.start_date,
            end_date=self.end_date,
            pdf_file=ContentFile(buffer.getvalue(), name=f'regulatory_{self.start_date}_{self.end_date}.pdf'),
            summary_data={
                'high_value_transactions': high_value_transactions.count(),
                'frequent_customers': frequent_customers.count(),
                'total_reported_volume': float(
                    high_value_transactions.aggregate(
                        total=Sum('amount_to')
                    )['total'] or 0
                )
            },
            generated_by=self.user
        )
        
        return report
    
    def _generate_regulatory_pdf(self, high_value_transactions, frequent_customers):
        """Genera PDF de reporte regulatorio"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
        story = []
        
        # Encabezado oficial
        story.append(Paragraph("REPORTE DE OPERACIONES CAMBIARIAS", self.styles['CustomTitle']))
        story.append(Paragraph("UNIDAD DE INVESTIGACIONES FINANCIERAS", self.styles['Heading2']))
        story.append(Paragraph(f"Período: {self.start_date} al {self.end_date}", self.styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Sección 1: Transacciones de alto valor
        story.append(Paragraph("1. TRANSACCIONES DE ALTO VALOR", self.styles['SectionTitle']))
        story.append(Paragraph("Operaciones superiores a USD 5,000 o equivalente", self.styles['Normal']))
        story.append(Spacer(1, 10))
        
        if high_value_transactions.exists():
            hv_data = [
                ['Fecha', 'Número', 'Cliente', 'Documento', 'Tipo', 'Divisa', 
                 'Monto', 'Tasa', 'Total BOB', 'Sucursal']
            ]
            
            for trans in high_value_transactions:
                hv_data.append([
                    trans.created_at.strftime('%Y-%m-%d %H:%M'),
                    trans.transaction_number,
                    trans.customer.full_name[:30],
                    f"{trans.customer.document_type} {trans.customer.document_number}",
                    trans.get_transaction_type_display(),
                    trans.currency_from.code,
                    f"{trans.amount_from:,.2f}",
                    f"{trans.exchange_rate:,.4f}",
                    f"{trans.amount_to:,.2f}",
                    trans.branch.name
                ])
            
            hv_table = Table(hv_data)
            hv_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
            ]))
            story.append(hv_table)
        else:
            story.append(Paragraph("No se registraron transacciones de alto valor en el período.", 
                                 self.styles['Normal']))
        
        story.append(PageBreak())
        
        # Sección 2: Clientes frecuentes
        story.append(Paragraph("2. CLIENTES FRECUENTES", self.styles['SectionTitle']))
        story.append(Paragraph("Clientes con 5 o más transacciones en el período", self.styles['Normal']))
        story.append(Spacer(1, 10))
        
        if frequent_customers:
            fc_data = [
                ['Nombre', 'Documento', 'Nacionalidad', 'PEP', 
                 'Transacciones', 'Volumen Total (BOB)']
            ]
            
            for customer in frequent_customers:
                fc_data.append([
                    customer['customer__full_name'][:40],
                    f"{customer['customer__document_type']} {customer['customer__document_number']}",
                    customer['customer__nationality'] or 'No especificada',
                    'Sí' if customer['customer__is_pep'] else 'No',
                    f"{customer['transaction_count']:,}",
                    f"Bs. {customer['total_volume_bob']:,.2f}"
               ])
           
            fc_table = Table(fc_data)
            fc_table.setStyle(TableStyle([
               ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
               ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
               ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
               ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
               ('FONTSIZE', (0, 0), (-1, 0), 9),
               ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
               ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
               ('GRID', (0, 0), (-1, -1), 1, colors.black),
               ('FONTSIZE', (0, 1), (-1, -1), 8),
           ]))
            story.append(fc_table)
        else:
            story.append(Paragraph("No se identificaron clientes frecuentes en el período.", 
                                self.styles['Normal']))
       
       # Declaración de veracidad
        story.append(Spacer(1, 50))
        story.append(Paragraph("DECLARACIÓN", self.styles['SectionTitle']))
        story.append(Paragraph(
           "El presente reporte ha sido elaborado de conformidad con las disposiciones "
           "legales vigentes y contiene información veraz y completa sobre las operaciones "
           "cambiarias realizadas en el período indicado.",
           self.styles['Normal']
       ))
       
        story.append(Spacer(1, 30))
        story.append(Paragraph("_" * 40, self.styles['Normal']))
        story.append(Paragraph("Firma del Oficial de Cumplimiento", self.styles['Normal']))
        story.append(Paragraph(f"Fecha: {datetime.now().strftime('%Y-%m-%d')}", self.styles['Normal']))
       
       # Construir PDF
        doc.build(story)
        buffer.seek(0)
       
        return buffer