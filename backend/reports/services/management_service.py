import os
import logging
from datetime import date, timedelta
from decimal import Decimal

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                  Paragraph, Spacer, HRFlowable)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER

from django.conf import settings
from django.db.models import Sum, Count, Avg, Q

logger = logging.getLogger(__name__)

CORP_DARK  = colors.HexColor('#1E3A5F')
CORP_LIGHT = colors.HexColor('#EBF3FB')
GREEN      = colors.HexColor('#1F7A4D')
RED        = colors.HexColor('#C0392B')
WHITE      = colors.white

REPORTS_DIR = os.path.join(settings.MEDIA_ROOT, 'reports', 'management')


def _ensure(path):
    os.makedirs(path, exist_ok=True)
    return path


def _xl_header(ws, row, headers):
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font      = Font(bold=True, color='FFFFFF', size=10)
        cell.fill      = PatternFill('solid', fgColor='1E3A5F')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = thin
    ws.row_dimensions[row].height = 18


def _pdf_doc(path):
    return SimpleDocTemplate(path, pagesize=A4,
                              topMargin=1.5*cm, bottomMargin=1.5*cm,
                              leftMargin=1.5*cm, rightMargin=1.5*cm)


def _pdf_styles():
    s      = getSampleStyleSheet()
    h1     = ParagraphStyle('H1', parent=s['Title'],    textColor=CORP_DARK, fontSize=16, spaceAfter=4)
    h2     = ParagraphStyle('H2', parent=s['Heading2'], textColor=CORP_DARK, fontSize=11, spaceBefore=10, spaceAfter=4)
    body   = ParagraphStyle('B',  parent=s['Normal'],   fontSize=9, leading=13)
    footer = ParagraphStyle('F',  parent=s['Normal'],   fontSize=7, textColor=colors.grey, alignment=TA_CENTER)
    sub    = ParagraphStyle('S',  parent=s['Normal'],   fontSize=9, textColor=colors.grey, alignment=TA_CENTER, spaceAfter=8)
    return h1, h2, body, footer, sub


def _pdf_ts():
    return TableStyle([
        ('BACKGROUND',     (0,0),(-1,0), CORP_DARK),
        ('TEXTCOLOR',      (0,0),(-1,0), WHITE),
        ('FONTNAME',       (0,0),(-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',       (0,0),(-1,0), 9),
        ('ALIGN',          (0,0),(-1,0), 'CENTER'),
        ('GRID',           (0,0),(-1,-1), 0.5, colors.HexColor('#DDDDDD')),
        ('FONTSIZE',       (0,1),(-1,-1), 8),
        ('ROWBACKGROUNDS', (0,1),(-1,-1), [WHITE, CORP_LIGHT]),
        ('BOTTOMPADDING',  (0,0),(-1,-1), 5),
        ('TOPPADDING',     (0,0),(-1,-1), 5),
        ('ALIGN',          (1,1),(-1,-1), 'RIGHT'),
    ])


class ManagementReportService:

    @classmethod
    def generate_pnl(cls, date_from, date_to, period='daily', user=None):
        from transactions.models import Transaction
        from reports.models import GeneratedReport
        from django.db.models.functions import TruncDate

        txs = Transaction.objects.filter(
            created_at__date__gte=date_from,
            created_at__date__lte=date_to,
            status='COMPLETED')
        daily = list(txs.annotate(day=TruncDate('created_at'))
                        .values('day')
                        .annotate(
                            buy=Sum('amount_to',  filter=Q(transaction_type='BUY')),
                            sell=Sum('amount_to', filter=Q(transaction_type='SELL')),
                            count=Count('id'))
                        .order_by('day'))

        tb  = float(txs.filter(transaction_type='BUY').aggregate(s=Sum('amount_to'))['s'] or 0)
        ts  = float(txs.filter(transaction_type='SELL').aggregate(s=Sum('amount_to'))['s'] or 0)
        tp  = ts - tb
        mgn = round((tp/ts*100) if ts else 0, 2)

        ep = cls._pnl_excel(daily, tb, ts, tp, mgn, date_from, date_to)
        pp = cls._pnl_pdf(daily, tb, ts, tp, mgn, date_from, date_to, period)

        if user:
            rtype = 'PNL_DAILY' if period == 'daily' else 'PNL_MONTHLY'
            for path, fmt in [(ep,'EXCEL'),(pp,'PDF')]:
                GeneratedReport.objects.create(
                    report_type=rtype, format=fmt,
                    date_from=date_from, date_to=date_to,
                    file_path=path.replace(settings.MEDIA_ROOT,'').lstrip('/'),
                    file_size_kb=os.path.getsize(path)//1024,
                    generated_by=user, parameters={'period':period})

        return dict(date_from=str(date_from), date_to=str(date_to),
                    total_buy_bob=tb, total_sell_bob=ts,
                    gross_profit_bob=tp, margin_pct=mgn,
                    total_transactions=txs.count(),
                    excel_path=ep, pdf_path=pp)

    @classmethod
    def _pnl_excel(cls, daily, tb, ts, tp, mgn, df, dt):
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'P&G'
        thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                      top=Side(style='thin'), bottom=Side(style='thin'))
        ca = Alignment(horizontal='center')
        ws.merge_cells('A1:F1')
        ws['A1'] = f'PERDIDAS Y GANANCIAS - {df.strftime("%d/%m/%Y")} al {dt.strftime("%d/%m/%Y")}'
        ws['A1'].font = Font(bold=True, size=14, color='1E3A5F'); ws['A1'].alignment = ca
        for col,(lbl,val,hex_) in enumerate([
            ('Total Compras', f'Bs. {tb:,.2f}',  '2E75B6'),
            ('Total Ventas',  f'Bs. {ts:,.2f}',  '1F7A4D'),
            ('Utilidad',      f'Bs. {tp:,.2f}',  '1F7A4D' if tp>=0 else 'C0392B'),
            ('Margen %',      f'{mgn}%',          '7B2D8B'),
        ], 1):
            ws.cell(row=3,column=col,value=lbl).font = Font(bold=True,color=hex_,size=10)
            ws.cell(row=4,column=col,value=val).font = Font(bold=True,size=12)
            ws.cell(row=3,column=col).alignment = ca
            ws.cell(row=4,column=col).alignment = ca
        _xl_header(ws,6,['Fecha','Compras (BOB)','Ventas (BOB)','Utilidad (BOB)','Margen %','Trans.'])
        for row_i,d in enumerate(daily,7):
            buy=float(d['buy'] or 0); sell=float(d['sell'] or 0)
            profit=sell-buy; margin=round((profit/sell*100) if sell else 0,2)
            for col,v in enumerate([d['day'].strftime('%d/%m/%Y'),f'{buy:,.2f}',
                                     f'{sell:,.2f}',f'{profit:,.2f}',f'{margin}%',d['count']],1):
                cell=ws.cell(row=row_i,column=col,value=v)
                cell.border=thin; cell.alignment=ca
                if col==4: cell.font=Font(bold=True,color='1F7A4D' if profit>=0 else 'C0392B')
        n=len(daily)
        if n>0:
            chart=BarChart(); chart.type='col'; chart.title='Ventas vs Compras'
            chart.style=10; chart.width=20; chart.height=12
            chart.add_data(Reference(ws,min_col=2,max_col=3,min_row=6,max_row=6+n),titles_from_data=True)
            chart.set_categories(Reference(ws,min_col=1,min_row=7,max_row=6+n))
            ws.add_chart(chart,'H6')
        for i,w in enumerate([14,16,16,16,10,14],1):
            ws.column_dimensions[get_column_letter(i)].width=w
        path=os.path.join(_ensure(REPORTS_DIR),
                          f'PnG_{df.strftime("%Y%m%d")}_{dt.strftime("%Y%m%d")}.xlsx')
        wb.save(path); return path

    @classmethod
    def _pnl_pdf(cls, daily, tb, ts, tp, mgn, df, dt, period):
        path=os.path.join(_ensure(REPORTS_DIR),
                          f'PnG_{df.strftime("%Y%m%d")}_{dt.strftime("%Y%m%d")}.pdf')
        doc=_pdf_doc(path); h1,h2,body,footer,sub=_pdf_styles(); story=[]
        story.append(Paragraph('INFORME DE PERDIDAS Y GANANCIAS',h1))
        story.append(Paragraph(
            f'{df.strftime("%d/%m/%Y")} - {dt.strftime("%d/%m/%Y")} | '
            f'{"Diario" if period=="daily" else "Mensual"}',sub))
        story.append(HRFlowable(width='100%',thickness=2,color=CORP_DARK))
        story.append(Spacer(1,0.3*cm))
        t_kpi=Table([['Total Compras','Total Ventas','Utilidad Bruta','Margen %'],
                     [f'Bs. {tb:,.2f}',f'Bs. {ts:,.2f}',f'Bs. {tp:,.2f}',f'{mgn}%']],
                    colWidths=[4.4*cm]*4)
        t_kpi.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0),CORP_DARK),('TEXTCOLOR',(0,0),(-1,0),WHITE),
            ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTNAME',(0,1),(-1,1),'Helvetica-Bold'),
            ('FONTSIZE',(0,0),(-1,0),9),('FONTSIZE',(0,1),(-1,1),11),
            ('ALIGN',(0,0),(-1,-1),'CENTER'),('TEXTCOLOR',(2,1),(2,1),GREEN if tp>=0 else RED),
            ('GRID',(0,0),(-1,-1),0.5,colors.HexColor('#DDDDDD')),
            ('BOTTOMPADDING',(0,0),(-1,-1),8),('TOPPADDING',(0,0),(-1,-1),8),
        ]))
        story.append(t_kpi); story.append(Spacer(1,0.4*cm))
        story.append(Paragraph('DETALLE POR DIA',h2))
        d_data=[['Fecha','Compras','Ventas','Utilidad','Margen %','Trans.']]
        for d in daily:
            buy=float(d['buy'] or 0); sell=float(d['sell'] or 0); p=sell-buy
            m=round((p/sell*100) if sell else 0,2)
            d_data.append([d['day'].strftime('%d/%m/%Y'),f'{buy:,.2f}',
                           f'{sell:,.2f}',f'{p:,.2f}',f'{m}%',str(d['count'])])
        t=Table(d_data,colWidths=[2.8*cm,3.2*cm,3.2*cm,3.2*cm,2*cm,1.6*cm],repeatRows=1)
        t.setStyle(_pdf_ts()); story.append(t)
        story.append(Spacer(1,1*cm))
        story.append(Paragraph(
            f'Forex ERP - Informe Gerencial - {date.today().strftime("%d/%m/%Y")}',footer))
        doc.build(story); return path

    @classmethod
    def generate_profitability(cls, date_from, date_to, user=None):
        from transactions.models import Transaction
        from reports.models import GeneratedReport

        txs=Transaction.objects.filter(
            created_at__date__gte=date_from,
            created_at__date__lte=date_to,status='COMPLETED')

        by_currency=[]
        for cur in txs.values_list('currency_from__code',flat=True).distinct():
            c=txs.filter(currency_from__code=cur)
            buy=float(c.filter(transaction_type='BUY').aggregate(s=Sum('amount_to'))['s'] or 0)
            sel=float(c.filter(transaction_type='SELL').aggregate(s=Sum('amount_to'))['s'] or 0)
            p=sel-buy
            by_currency.append(dict(currency=cur,buy_bob=buy,sell_bob=sel,profit=p,
                                    margin_pct=round((p/sel*100) if sel else 0,2),
                                    count=c.count()))

        by_branch=[]
        for b in txs.values('branch__name','branch_id').distinct():
            bt=txs.filter(branch_id=b['branch_id'])
            buy=float(bt.filter(transaction_type='BUY').aggregate(s=Sum('amount_to'))['s'] or 0)
            sel=float(bt.filter(transaction_type='SELL').aggregate(s=Sum('amount_to'))['s'] or 0)
            by_branch.append(dict(branch=b['branch__name'],buy_bob=buy,
                                  sell_bob=sel,profit=sel-buy,count=bt.count()))

        ep=cls._profit_excel(by_currency,by_branch,date_from,date_to)
        pp=cls._profit_pdf(by_currency,by_branch,date_from,date_to)
        if user:
            for path,fmt in [(ep,'EXCEL'),(pp,'PDF')]:
                GeneratedReport.objects.create(
                    report_type='PROFITABILITY',format=fmt,
                    date_from=date_from,date_to=date_to,
                    file_path=path.replace(settings.MEDIA_ROOT,'').lstrip('/'),
                    file_size_kb=os.path.getsize(path)//1024,generated_by=user)
        return dict(by_currency=by_currency,by_branch=by_branch,
                    excel_path=ep,pdf_path=pp)

    @classmethod
    def _profit_excel(cls,by_currency,by_branch,df,dt):
        wb=openpyxl.Workbook()
        thin=Border(left=Side(style='thin'),right=Side(style='thin'),
                    top=Side(style='thin'),bottom=Side(style='thin'))
        ca=Alignment(horizontal='center')
        ws1=wb.active; ws1.title='Por Divisa'
        ws1.merge_cells('A1:F1')
        ws1['A1']=f'RENTABILIDAD POR DIVISA - {df.strftime("%d/%m/%Y")} al {dt.strftime("%d/%m/%Y")}'
        ws1['A1'].font=Font(bold=True,size=13,color='1E3A5F')
        _xl_header(ws1,3,['Divisa','Compras (BOB)','Ventas (BOB)','Utilidad','Margen %','Trans.'])
        for i,r in enumerate(sorted(by_currency,key=lambda x:-x['profit']),4):
            for col,v in enumerate([r['currency'],f"{r['buy_bob']:,.2f}",
                                     f"{r['sell_bob']:,.2f}",f"{r['profit']:,.2f}",
                                     f"{r['margin_pct']}%",r['count']],1):
                cell=ws1.cell(row=i,column=col,value=v); cell.border=thin; cell.alignment=ca
                if col==4: cell.font=Font(bold=True,color='1F7A4D' if r['profit']>=0 else 'C0392B')
        for i,w in enumerate([10,18,18,18,12,10],1):
            ws1.column_dimensions[get_column_letter(i)].width=w
        ws2=wb.create_sheet('Por Sucursal')
        ws2.merge_cells('A1:E1')
        ws2['A1']='RENTABILIDAD POR SUCURSAL'
        ws2['A1'].font=Font(bold=True,size=13,color='1E3A5F')
        _xl_header(ws2,3,['Sucursal','Compras (BOB)','Ventas (BOB)','Utilidad','Trans.'])
        for i,r in enumerate(sorted(by_branch,key=lambda x:-x['profit']),4):
            for col,v in enumerate([r['branch'],f"{r['buy_bob']:,.2f}",
                                     f"{r['sell_bob']:,.2f}",f"{r['profit']:,.2f}",r['count']],1):
                cell=ws2.cell(row=i,column=col,value=v); cell.border=thin; cell.alignment=ca
        path=os.path.join(_ensure(REPORTS_DIR),
                          f'Rentabilidad_{df.strftime("%Y%m%d")}_{dt.strftime("%Y%m%d")}.xlsx')
        wb.save(path); return path

    @classmethod
    def _profit_pdf(cls,by_currency,by_branch,df,dt):
        path=os.path.join(_ensure(REPORTS_DIR),
                          f'Rentabilidad_{df.strftime("%Y%m%d")}_{dt.strftime("%Y%m%d")}.pdf')
        doc=_pdf_doc(path); h1,h2,body,footer,sub=_pdf_styles(); story=[]
        story.append(Paragraph('RENTABILIDAD POR DIVISA Y SUCURSAL',h1))
        story.append(Paragraph(f'{df.strftime("%d/%m/%Y")} - {dt.strftime("%d/%m/%Y")}',sub))
        story.append(HRFlowable(width='100%',thickness=2,color=CORP_DARK))
        story.append(Spacer(1,0.3*cm))
        story.append(Paragraph('POR DIVISA',h2))
        t1=Table([['Divisa','Compras','Ventas','Utilidad','Margen %','Trans.']]+
                 [[r['currency'],f"{r['buy_bob']:,.2f}",f"{r['sell_bob']:,.2f}",
                   f"{r['profit']:,.2f}",f"{r['margin_pct']}%",r['count']]
                  for r in sorted(by_currency,key=lambda x:-x['profit'])],
                 colWidths=[2*cm,3.2*cm,3.2*cm,3.2*cm,2.2*cm,1.8*cm],repeatRows=1)
        t1.setStyle(_pdf_ts()); story.append(t1); story.append(Spacer(1,0.4*cm))
        story.append(Paragraph('POR SUCURSAL',h2))
        t2=Table([['Sucursal','Compras','Ventas','Utilidad','Trans.']]+
                 [[r['branch'],f"{r['buy_bob']:,.2f}",f"{r['sell_bob']:,.2f}",
                   f"{r['profit']:,.2f}",r['count']]
                  for r in sorted(by_branch,key=lambda x:-x['profit'])],
                 colWidths=[5*cm,3.2*cm,3.2*cm,3.2*cm,1.8*cm],repeatRows=1)
        t2.setStyle(_pdf_ts()); story.append(t2)
        story.append(Spacer(1,1*cm))
        story.append(Paragraph(
            f'Forex ERP - Informe Gerencial - {date.today().strftime("%d/%m/%Y")}',footer))
        doc.build(story); return path

    @classmethod
    def generate_client_ranking(cls, date_from, date_to, top_n=20, user=None):
        from transactions.models import Transaction
        from reports.models import GeneratedReport

        data=list(Transaction.objects
                  .filter(created_at__date__gte=date_from,
                          created_at__date__lte=date_to,status='COMPLETED')
                  .values('customer__full_name','customer__document_number',
                          'customer__is_pep','customer_id')
                  .annotate(total_volume=Sum('amount_to'),
                            total_transactions=Count('id'),
                            avg_amount=Avg('amount_to'))
                  .order_by('-total_volume')[:top_n])

        ep=cls._ranking_excel(data,date_from,date_to,top_n)
        pp=cls._ranking_pdf(data,date_from,date_to,top_n)
        if user:
            for path,fmt in [(ep,'EXCEL'),(pp,'PDF')]:
                GeneratedReport.objects.create(
                    report_type='CLIENT_RANKING',format=fmt,
                    date_from=date_from,date_to=date_to,
                    file_path=path.replace(settings.MEDIA_ROOT,'').lstrip('/'),
                    file_size_kb=os.path.getsize(path)//1024,
                    generated_by=user,parameters={'top_n':top_n})
        return dict(data=data,excel_path=ep,pdf_path=pp)

    @classmethod
    def _ranking_excel(cls,data,df,dt,top_n):
        wb=openpyxl.Workbook(); ws=wb.active; ws.title=f'Top {top_n}'
        thin=Border(left=Side(style='thin'),right=Side(style='thin'),
                    top=Side(style='thin'),bottom=Side(style='thin'))
        ca=Alignment(horizontal='center')
        ws.merge_cells('A1:F1')
        ws['A1']=f'RANKING TOP {top_n} CLIENTES - {df.strftime("%d/%m/%Y")} al {dt.strftime("%d/%m/%Y")}'
        ws['A1'].font=Font(bold=True,size=13,color='1E3A5F')
        _xl_header(ws,3,['#','Cliente','Documento','Volumen (BOB)','Trans.','Prom. Trans.'])
        for i,r in enumerate(data,1):
            pep=' (PEP)' if r['customer__is_pep'] else ''
            for col,v in enumerate([i,r['customer__full_name']+pep,
                                     r['customer__document_number'],
                                     f"{float(r['total_volume'] or 0):,.2f}",
                                     r['total_transactions'],
                                     f"{float(r['avg_amount'] or 0):,.2f}"],1):
                cell=ws.cell(row=i+3,column=col,value=v); cell.border=thin; cell.alignment=ca
                if col==4: cell.font=Font(bold=True,color='1E3A5F')
            if r['customer__is_pep']:
                for col in range(1,7):
                    ws.cell(row=i+3,column=col).fill=PatternFill('solid',fgColor='FFE0E0')
        for i,w in enumerate([5,30,18,18,10,16],1):
            ws.column_dimensions[get_column_letter(i)].width=w
        path=os.path.join(_ensure(REPORTS_DIR),
                          f'RankingClientes_{df.strftime("%Y%m%d")}_{dt.strftime("%Y%m%d")}.xlsx')
        wb.save(path); return path

    @classmethod
    def _ranking_pdf(cls,data,df,dt,top_n):
        path=os.path.join(_ensure(REPORTS_DIR),
                          f'RankingClientes_{df.strftime("%Y%m%d")}_{dt.strftime("%Y%m%d")}.pdf')
        doc=_pdf_doc(path); h1,h2,body,footer,sub=_pdf_styles(); story=[]
        story.append(Paragraph(f'RANKING TOP {top_n} CLIENTES',h1))
        story.append(Paragraph(f'{df.strftime("%d/%m/%Y")} - {dt.strftime("%d/%m/%Y")}',sub))
        story.append(HRFlowable(width='100%',thickness=2,color=CORP_DARK))
        story.append(Spacer(1,0.3*cm))
        t_data=[['#','Cliente','Documento','Volumen (BOB)','Trans.','Prom.']]
        for i,r in enumerate(data,1):
            pep=' *' if r['customer__is_pep'] else ''
            t_data.append([str(i),r['customer__full_name'][:28]+pep,
                           r['customer__document_number'],
                           f"{float(r['total_volume'] or 0):,.2f}",
                           r['total_transactions'],
                           f"{float(r['avg_amount'] or 0):,.2f}"])
        t=Table(t_data,colWidths=[1*cm,5.5*cm,3.5*cm,3.5*cm,1.5*cm,3*cm],repeatRows=1)
        t.setStyle(_pdf_ts()); story.append(t)
        story.append(Spacer(1,0.3*cm))
        story.append(Paragraph('* Cliente PEP (Persona Expuesta Politicamente)',body))
        story.append(Spacer(1,1*cm))
        story.append(Paragraph(
            f'Forex ERP - Informe Gerencial - {date.today().strftime("%d/%m/%Y")}',footer))
        doc.build(story); return path

    @classmethod
    def generate_comparative(cls, date_from, date_to, user=None):
        from transactions.models import Transaction
        from reports.models import GeneratedReport

        delta=date_to-date_from
        pf=date_from-(delta+timedelta(days=1)); pt=date_from-timedelta(days=1)

        def _t(df,dt):
            qs=Transaction.objects.filter(
                created_at__date__gte=df,created_at__date__lte=dt,status='COMPLETED')
            buy=float(qs.filter(transaction_type='BUY').aggregate(s=Sum('amount_to'))['s'] or 0)
            sell=float(qs.filter(transaction_type='SELL').aggregate(s=Sum('amount_to'))['s'] or 0)
            return dict(buy=buy,sell=sell,profit=sell-buy,count=qs.count())

        curr=_t(date_from,date_to); prev=_t(pf,pt)
        chg=dict(
            buy_pct=   round(((curr['buy']   -prev['buy'])   /prev['buy']   *100) if prev['buy']    else 0,2),
            sell_pct=  round(((curr['sell']  -prev['sell'])  /prev['sell']  *100) if prev['sell']   else 0,2),
            profit_pct=round(((curr['profit']-prev['profit'])/prev['profit']*100) if prev['profit'] else 0,2),
            count_pct= round(((curr['count'] -prev['count']) /prev['count'] *100) if prev['count']  else 0,2),
        )
        ep=cls._comparative_excel(curr,prev,chg,date_from,date_to,pf,pt)
        pp=cls._comparative_pdf(curr,prev,chg,date_from,date_to,pf,pt)
        if user:
            for path,fmt in [(ep,'EXCEL'),(pp,'PDF')]:
                GeneratedReport.objects.create(
                    report_type='COMPARATIVE',format=fmt,
                    date_from=date_from,date_to=date_to,
                    file_path=path.replace(settings.MEDIA_ROOT,'').lstrip('/'),
                    file_size_kb=os.path.getsize(path)//1024,generated_by=user)
        return dict(current=curr,previous=prev,changes=chg,excel_path=ep,pdf_path=pp)

    @classmethod
    def _comparative_excel(cls,curr,prev,chg,df,dt,pf,pt):
        wb=openpyxl.Workbook(); ws=wb.active; ws.title='Comparativo'
        thin=Border(left=Side(style='thin'),right=Side(style='thin'),
                    top=Side(style='thin'),bottom=Side(style='thin'))
        ca=Alignment(horizontal='center',vertical='center')
        ws.merge_cells('A1:E1')
        ws['A1']=f'COMPARATIVO DE PERIODOS - {df} vs {pf}'
        ws['A1'].font=Font(bold=True,size=13,color='1E3A5F'); ws['A1'].alignment=ca
        _xl_header(ws,3,['Indicador',f'Actual ({df}>{dt})',
                          f'Anterior ({pf}>{pt})','Variacion (BOB)','Cambio %'])
        for row_i,(lbl,cv,pv,pct) in enumerate([
            ('Compras (BOB)', curr['buy'],    prev['buy'],    chg['buy_pct']),
            ('Ventas (BOB)',  curr['sell'],   prev['sell'],   chg['sell_pct']),
            ('Utilidad (BOB)',curr['profit'], prev['profit'], chg['profit_pct']),
            ('Transacciones', curr['count'],  prev['count'],  chg['count_pct']),
        ],4):
            diff=cv-pv
            for col,v in enumerate([lbl,f'{float(cv):,.2f}',f'{float(pv):,.2f}',
                                     f'{float(diff):+,.2f}',f'{pct:+.2f}%'],1):
                cell=ws.cell(row=row_i,column=col,value=v); cell.border=thin; cell.alignment=ca
                if col in (4,5): cell.font=Font(bold=True,color='1F7A4D' if diff>=0 else 'C0392B')
        for i,w in enumerate([25,22,22,18,12],1):
            ws.column_dimensions[get_column_letter(i)].width=w
        path=os.path.join(_ensure(REPORTS_DIR),f'Comparativo_{df.strftime("%Y%m%d")}.xlsx')
        wb.save(path); return path

    @classmethod
    def _comparative_pdf(cls,curr,prev,chg,df,dt,pf,pt):
        path=os.path.join(_ensure(REPORTS_DIR),f'Comparativo_{df.strftime("%Y%m%d")}.pdf')
        doc=_pdf_doc(path); h1,h2,body,footer,sub=_pdf_styles(); story=[]
        story.append(Paragraph('COMPARATIVO DE PERIODOS',h1))
        story.append(Paragraph(f'Actual: {df} - {dt} | Anterior: {pf} - {pt}',sub))
        story.append(HRFlowable(width='100%',thickness=2,color=CORP_DARK))
        story.append(Spacer(1,0.3*cm))
        t_data=[
            ['Indicador',f'Actual\n{df}',f'Anterior\n{pf}','Variacion','Cambio %'],
            ['Compras',  f"{curr['buy']:,.2f}",   f"{prev['buy']:,.2f}",
             f"{curr['buy']-prev['buy']:+,.2f}",   f"{chg['buy_pct']:+.2f}%"],
            ['Ventas',   f"{curr['sell']:,.2f}",  f"{prev['sell']:,.2f}",
             f"{curr['sell']-prev['sell']:+,.2f}", f"{chg['sell_pct']:+.2f}%"],
            ['Utilidad', f"{curr['profit']:,.2f}",f"{prev['profit']:,.2f}",
             f"{curr['profit']-prev['profit']:+,.2f}",f"{chg['profit_pct']:+.2f}%"],
            ['Trans.',   str(curr['count']),str(prev['count']),
             f"{curr['count']-prev['count']:+d}", f"{chg['count_pct']:+.2f}%"],
        ]
        t=Table(t_data,colWidths=[3.5*cm,3.5*cm,3.5*cm,3*cm,2.5*cm],repeatRows=1)
        ts=_pdf_ts()
        for ri in range(1,len(t_data)):
            try:
                diff=float(str(t_data[ri][3]).replace(',','').replace('+',''))
                col_=GREEN if diff>=0 else RED
                ts.add('TEXTCOLOR',(3,ri),(4,ri),col_)
                ts.add('FONTNAME',(3,ri),(4,ri),'Helvetica-Bold')
            except (ValueError,IndexError): pass
        t.setStyle(ts); story.append(t)
        story.append(Spacer(1,1*cm))
        story.append(Paragraph(
            f'Forex ERP - Informe Gerencial - {date.today().strftime("%d/%m/%Y")}',footer))
        doc.build(story); return path

    @classmethod
    def generate_cashflow_projection(cls, base_date, days_ahead=30, user=None):
        from transactions.models import Transaction
        from django.db.models.functions import TruncDate
        from reports.models import GeneratedReport

        hist=list(Transaction.objects
                  .filter(created_at__date__gte=base_date-timedelta(days=60),
                          created_at__date__lte=base_date,status='COMPLETED')
                  .annotate(day=TruncDate('created_at')).values('day')
                  .annotate(buy=Sum('amount_to',filter=Q(transaction_type='BUY')),
                            sell=Sum('amount_to',filter=Q(transaction_type='SELL')))
                  .order_by('day'))
        hist=[dict(date=str(d['day']),buy=float(d['buy'] or 0),sell=float(d['sell'] or 0),
                   profit=float(d['sell'] or 0)-float(d['buy'] or 0)) for d in hist]

        if hist:
            avg_buy=sum(d['buy'] for d in hist)/len(hist)
            avg_sell=sum(d['sell'] for d in hist)/len(hist)
            avg_p=sum(d['profit'] for d in hist)/len(hist)
            rec=hist[-14:] if len(hist)>=14 else hist
            trend=(rec[-1]['profit']-rec[0]['profit'])/len(rec) if len(rec)>1 else 0
        else:
            avg_buy=avg_sell=avg_p=trend=0

        projection=[]
        for i in range(1,days_ahead+1):
            d=base_date+timedelta(days=i)
            projection.append(dict(
                date=str(d),
                projected_buy=round(avg_buy,2),
                projected_sell=round(avg_sell+avg_sell*(trend/(avg_sell or 1))*i*0.1,2),
                projected_profit=round(avg_p+trend*i,2),
                is_weekend=d.weekday()>=5))

        ep=cls._cashflow_excel(projection,base_date,days_ahead)
        pp=cls._cashflow_pdf(projection,base_date,days_ahead,avg_p,trend)
        if user:
            for path,fmt in [(ep,'EXCEL'),(pp,'PDF')]:
                GeneratedReport.objects.create(
                    report_type='CASHFLOW',format=fmt,
                    date_from=base_date,
                    date_to=base_date+timedelta(days=days_ahead),
                    file_path=path.replace(settings.MEDIA_ROOT,'').lstrip('/'),
                    file_size_kb=os.path.getsize(path)//1024,
                    generated_by=user,parameters={'days_ahead':days_ahead})
        return dict(base_date=str(base_date),avg_daily_profit=round(avg_p,2),
                    trend_daily=round(trend,2),projection=projection,
                    excel_path=ep,pdf_path=pp)

    @classmethod
    def _cashflow_excel(cls,projection,base_date,days_ahead):
        wb=openpyxl.Workbook(); ws=wb.active; ws.title='Flujo de Caja'
        thin=Border(left=Side(style='thin'),right=Side(style='thin'),
                    top=Side(style='thin'),bottom=Side(style='thin'))
        ca=Alignment(horizontal='center')
        ws.merge_cells('A1:E1')
        ws['A1']=f'PROYECCION DE FLUJO DE CAJA - Base: {base_date} + {days_ahead} dias'
        ws['A1'].font=Font(bold=True,size=13,color='1E3A5F')
        _xl_header(ws,3,['Fecha','Compras Proy.','Ventas Proy.','Utilidad Proy.','Fin de semana'])
        for i,r in enumerate(projection,4):
            for col,v in enumerate([r['date'],f"{r['projected_buy']:,.2f}",
                                     f"{r['projected_sell']:,.2f}",
                                     f"{r['projected_profit']:,.2f}",
                                     'Si' if r['is_weekend'] else ''],1):
                cell=ws.cell(row=i,column=col,value=v); cell.border=thin; cell.alignment=ca
                if col==4: cell.font=Font(bold=True,
                                         color='1F7A4D' if r['projected_profit']>=0 else 'C0392B')
            if r['is_weekend']:
                for col in range(1,6):
                    ws.cell(row=i,column=col).fill=PatternFill('solid',fgColor='F5F5F5')
        for i,w in enumerate([14,18,18,18,14],1):
            ws.column_dimensions[get_column_letter(i)].width=w
        path=os.path.join(_ensure(REPORTS_DIR),
                          f'FlujoCaja_{base_date.strftime("%Y%m%d")}.xlsx')
        wb.save(path); return path

    @classmethod
    def _cashflow_pdf(cls,projection,base_date,days_ahead,avg_p,trend):
        path=os.path.join(_ensure(REPORTS_DIR),
                          f'FlujoCaja_{base_date.strftime("%Y%m%d")}.pdf')
        doc=_pdf_doc(path); h1,h2,body,footer,sub=_pdf_styles(); story=[]
        story.append(Paragraph('PROYECCION DE FLUJO DE CAJA',h1))
        story.append(Paragraph(
            f'Base: {base_date} | Horizonte: {days_ahead} dias | '
            f'Prom. diario: Bs. {avg_p:,.2f} | Tendencia: {trend:+.2f}/dia',sub))
        story.append(HRFlowable(width='100%',thickness=2,color=CORP_DARK))
        story.append(Spacer(1,0.3*cm))
        t_data=[['Fecha','Compras Proy.','Ventas Proy.','Utilidad Proy.']]
        for r in projection[:20]:
            t_data.append([r['date'],f"{r['projected_buy']:,.2f}",
                           f"{r['projected_sell']:,.2f}",f"{r['projected_profit']:,.2f}"])
        t=Table(t_data,colWidths=[3.5*cm,4*cm,4*cm,4*cm],repeatRows=1)
        t.setStyle(_pdf_ts()); story.append(t)
        if len(projection)>20:
            story.append(Paragraph(
                f'... y {len(projection)-20} dias mas en el archivo Excel.',body))
        story.append(Spacer(1,1*cm))
        story.append(Paragraph(
            f'Proyeccion basada en promedio historico 60 dias. '
            f'Forex ERP - {date.today().strftime("%d/%m/%Y")}',footer))
        doc.build(story); return path