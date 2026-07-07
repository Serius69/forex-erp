import os
import logging
from datetime import date, datetime
from decimal import Decimal

# Logger separado para auditoría de filtrado ASFI — va a kapitalya.asfi_audit
audit_log = logging.getLogger('kapitalya.asfi_audit')

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                  Paragraph, Spacer, HRFlowable)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER

from django.conf import settings
from django.db import models
from django.db.models import Sum, Count, Q

logger = logging.getLogger(__name__)

ASFI_DARK  = colors.HexColor('#003366')
ASFI_LIGHT = colors.HexColor('#CCE0FF')
RED_CONF   = colors.HexColor('#CC0000')
WHITE      = colors.white
GRAY_ROW   = colors.HexColor('#F5F5F5')

REPORTS_DIR = os.path.join(settings.MEDIA_ROOT, 'reports', 'asfi')


def _ensure(path):
    os.makedirs(path, exist_ok=True)
    return path


def _xl_header(ws, row, headers):
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font      = Font(bold=True, color='FFFFFF', size=9)
        cell.fill      = PatternFill('solid', fgColor='003366')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = thin
    ws.row_dimensions[row].height = 18


def _pdf_table_style():
    return TableStyle([
        ('BACKGROUND',     (0, 0), (-1, 0), ASFI_DARK),
        ('TEXTCOLOR',      (0, 0), (-1, 0), WHITE),
        ('FONTNAME',       (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',       (0, 0), (-1, 0), 8),
        ('ALIGN',          (0, 0), (-1, 0), 'CENTER'),
        ('GRID',           (0, 0), (-1,-1), 0.5, colors.HexColor('#CCCCCC')),
        ('FONTSIZE',       (0, 1), (-1,-1), 7),
        ('ROWBACKGROUNDS', (0, 1), (-1,-1), [WHITE, GRAY_ROW]),
        ('ALIGN',          (0, 1), (-1,-1), 'CENTER'),
        ('BOTTOMPADDING',  (0, 0), (-1,-1), 4),
        ('TOPPADDING',     (0, 0), (-1,-1), 4),
    ])


class ASFIReportService:

    # ── RTE Mensual ───────────────────────────────────────────────────────────

    @classmethod
    def generate_rte_monthly(cls, year: int, month: int, user=None) -> dict:
        from reports.models import CashTransactionReport, GeneratedReport
        import calendar

        start = date(year, month, 1)
        end   = date(year, month, calendar.monthrange(year, month)[1])

        # ── Conteo total (sin filtro) para calcular excluidas ─────────────────
        total_universe = CashTransactionReport.objects.filter(
            report_date__gte=start, report_date__lte=end,
        ).count()

        # ── Queryset filtrado: SOLO transacciones visibles para ASFI ──────────
        # visible_asfi=True es el campo canónico; is_reportable_to_asfi se
        # mantiene como alias en sync dentro de Transaction.save().
        rtes = (CashTransactionReport.objects
                .filter(report_date__gte=start, report_date__lte=end,
                        transaction__visible_asfi=True)
                .select_related('transaction', 'transaction__customer')
                .order_by('report_date'))

        total_included = rtes.count()
        total_excluded = total_universe - total_included

        # ── Validación pre-generación: ninguna transacción interna puede ──────
        # haber pasado el filtro (defensa en profundidad sobre el ORM filter).
        leaked = rtes.filter(transaction__visible_asfi=False).count()
        if leaked > 0:
            audit_log.critical(
                'ASFI_RTE_VALIDATION_FAIL year=%d month=%d leaked_internal=%d '
                'user=%s — reporte NO generado',
                year, month, leaked, getattr(user, 'username', 'system'),
            )
            raise ValueError(
                f'Validación ASFI fallida: {leaked} transacción(es) interna(s) '
                f'detectada(s) en el queryset RTE {year}-{month:02d}. '
                f'El reporte no fue generado.'
            )

        # ── Log de auditoría ──────────────────────────────────────────────────
        audit_log.info(
            'ASFI_RTE_GENERATE year=%d month=%d included=%d excluded_internal=%d '
            'user=%s',
            year, month, total_included, total_excluded,
            getattr(user, 'username', 'system'),
        )

        excel_path = cls._rte_excel(rtes, start, end, year, month)
        pdf_path   = cls._rte_pdf(rtes, start, end, year, month)

        if user:
            for path, fmt in [(excel_path, 'EXCEL'), (pdf_path, 'PDF')]:
                GeneratedReport.objects.create(
                    report_type='RTE_MONTHLY', format=fmt,
                    date_from=start, date_to=end,
                    file_path=path.replace(settings.MEDIA_ROOT, '').lstrip('/'),
                    file_size_kb=os.path.getsize(path) // 1024,
                    generated_by=user,
                    parameters={
                        'year': year, 'month': month,
                        'total_included': total_included,
                        'total_excluded': total_excluded,
                    },
                )
        return {
            'status':           'ok',
            'year':             year,
            'month':            month,
            'total_records':    total_included,
            'total_excluded':   total_excluded,
            'excel_path':       excel_path,
            'pdf_path':         pdf_path,
        }

    @classmethod
    def _rte_excel(cls, rtes, start, end, year, month) -> str:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f'RTE {year}-{month:02d}'
        thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                      top=Side(style='thin'), bottom=Side(style='thin'))
        ca = Alignment(horizontal='center', vertical='center')

        for merge, text, sz in [
            ('A1:L1', 'AUTORIDAD DE SUPERVISION DEL SISTEMA FINANCIERO - ASFI', 13),
            ('A2:L2', f'REGISTRO DE TRANSACCIONES EN EFECTIVO (RTE) - {start.strftime("%B %Y").upper()}', 11),
            ('A3:L3', f'Casa de Cambio Forex ERP | {start.strftime("%d/%m/%Y")} al {end.strftime("%d/%m/%Y")}', 10),
        ]:
            ws.merge_cells(merge)
            cell = ws[merge.split(':')[0]]
            cell.value = text
            cell.font  = Font(bold=True, size=sz, color='003366')
            cell.alignment = ca

        _xl_header(ws, 5, ['N RTE','Fecha','N Transaccion','Tipo',
                             'Cliente','Tipo Doc.','N Documento','Nac.',
                             'Monto Original','Divisa','Equiv. USD','PEP'])

        total_usd = Decimal('0')
        for row_i, rte in enumerate(rtes, 6):
            tx       = rte.transaction
            pep_fill = PatternFill('solid', fgColor='FFE0E0') if rte.customer_is_pep else None
            vals = [
                rte.report_number,
                rte.report_date.strftime('%d/%m/%Y'),
                tx.transaction_number,
                'COMPRA' if tx.transaction_type == 'BUY' else 'VENTA',
                rte.customer_full_name,
                rte.customer_document_type,
                rte.customer_document_num,
                rte.customer_nationality[:3].upper(),
                f'{float(rte.original_amount):,.2f}',
                rte.currency_code,
                f'{float(rte.amount_usd_equiv):,.2f}',
                'SI' if rte.customer_is_pep else 'NO',
            ]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row_i, column=col, value=val)
                cell.border = thin; cell.alignment = ca
                if col == 11: cell.font = Font(bold=True)
                if pep_fill:  cell.fill = pep_fill
            total_usd += rte.amount_usd_equiv

        tr = rtes.count() + 6
        ws.merge_cells(f'A{tr}:J{tr}')
        ws[f'A{tr}']      = f'TOTAL: {rtes.count()} registros'
        ws[f'A{tr}'].font = Font(bold=True)
        ws[f'K{tr}']      = f'${float(total_usd):,.2f}'
        ws[f'K{tr}'].font = Font(bold=True, color='003366')

        for i, w in enumerate([14,12,18,8,30,10,18,6,14,8,14,6], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        path = os.path.join(_ensure(REPORTS_DIR), f'RTE_{year}{month:02d}.xlsx')
        wb.save(path)
        return path

    @classmethod
    def _rte_pdf(cls, rtes, start, end, year, month) -> str:
        path   = os.path.join(_ensure(REPORTS_DIR), f'RTE_{year}{month:02d}.pdf')
        doc    = SimpleDocTemplate(path, pagesize=A4,
                                    topMargin=1.5*cm, bottomMargin=1.5*cm,
                                    leftMargin=1.2*cm, rightMargin=1.2*cm)
        styles = getSampleStyleSheet()
        story  = []

        story.append(Paragraph('AUTORIDAD DE SUPERVISION DEL SISTEMA FINANCIERO',
            ParagraphStyle('T1', parent=styles['Title'], textColor=ASFI_DARK, fontSize=13)))
        story.append(Paragraph(
            f'RTE - {start.strftime("%B %Y").upper()}',
            ParagraphStyle('T2', parent=styles['Normal'], textColor=ASFI_DARK,
                           fontSize=10, alignment=TA_CENTER, spaceAfter=2)))
        story.append(Paragraph(
            f'Forex ERP | {start.strftime("%d/%m/%Y")} - {end.strftime("%d/%m/%Y")}',
            ParagraphStyle('S', parent=styles['Normal'], fontSize=8,
                           textColor=colors.grey, alignment=TA_CENTER, spaceAfter=8)))
        story.append(HRFlowable(width='100%', thickness=1.5, color=ASFI_DARK))
        story.append(Spacer(1, 0.3*cm))

        headers   = ['N RTE','Fecha','N Trans.','Tipo','Cliente',
                     'Documento','Monto','Divisa','USD Equiv.','PEP']
        data      = [headers]
        total_usd = Decimal('0')

        for rte in rtes:
            data.append([
                rte.report_number,
                rte.report_date.strftime('%d/%m/%Y'),
                rte.transaction.transaction_number,
                'BUY' if rte.transaction.transaction_type == 'BUY' else 'SELL',
                rte.customer_full_name[:22],
                f'{rte.customer_document_type}: {rte.customer_document_num}',
                f'{float(rte.original_amount):,.2f}',
                rte.currency_code,
                f'${float(rte.amount_usd_equiv):,.2f}',
                'PEP' if rte.customer_is_pep else '-',
            ])
            total_usd += rte.amount_usd_equiv

        data.append(['','','','', f'TOTAL: {rtes.count()}','','','',
                     f'${float(total_usd):,.2f}',''])

        t = Table(data,
                  colWidths=[2.2*cm,1.8*cm,2.5*cm,1.2*cm,
                              3.8*cm,3.2*cm,2*cm,1.2*cm,2.2*cm,1.2*cm],
                  repeatRows=1)
        ts = _pdf_table_style()
        ts.add('FONTNAME',   (0,-1),(-1,-1), 'Helvetica-Bold')
        ts.add('BACKGROUND', (0,-1),(-1,-1), ASFI_LIGHT)
        t.setStyle(ts)
        story.append(t)
        story.append(Spacer(1, 0.5*cm))
        story.append(Paragraph(
            f'Generado el {datetime.now().strftime("%d/%m/%Y %H:%M")} - DOCUMENTO CONFIDENCIAL ASFI',
            ParagraphStyle('F', parent=styles['Normal'],
                           fontSize=7, textColor=colors.grey, alignment=TA_CENTER)))
        doc.build(story)
        return path

    # ── ROUE ──────────────────────────────────────────────────────────────────

    @classmethod
    def generate_roue_pdf(cls, sar_id: int, user=None) -> str:
        from reports.models import SuspiciousActivityReport, GeneratedReport

        sar    = (SuspiciousActivityReport.objects
                  .select_related('customer', 'detected_by')
                  .prefetch_related('transactions')
                  .get(pk=sar_id))
        path   = os.path.join(_ensure(REPORTS_DIR), f'ROUE_{sar.report_number}.pdf')
        doc    = SimpleDocTemplate(path, pagesize=A4,
                                    topMargin=2*cm, bottomMargin=2*cm,
                                    leftMargin=2*cm, rightMargin=2*cm)
        styles = getSampleStyleSheet()
        story  = []

        risk_hex = {'LOW':'#28A745','MEDIUM':'#FFC107','HIGH':'#DC3545','CRITICAL':'#7B0000'}

        story.append(Paragraph('AUTORIDAD DE SUPERVISION DEL SISTEMA FINANCIERO',
            ParagraphStyle('T1', parent=styles['Title'], textColor=ASFI_DARK, fontSize=14)))
        story.append(Paragraph('REPORTE DE OPERACION INUSUAL O SOSPECHOSA (ROUE)',
            ParagraphStyle('T2', parent=styles['Heading2'], textColor=ASFI_DARK, fontSize=11)))
        story.append(HRFlowable(width='100%', thickness=2, color=ASFI_DARK))
        story.append(Spacer(1, 0.3*cm))
        story.append(Paragraph(f'NIVEL DE RIESGO: {sar.risk_level}',
            ParagraphStyle('R', parent=styles['Normal'], fontSize=12,
                           fontName='Helvetica-Bold',
                           textColor=colors.HexColor(risk_hex.get(sar.risk_level,'#333')),
                           alignment=TA_CENTER)))
        story.append(Spacer(1, 0.3*cm))

        meta = [
            ['N Reporte:', sar.report_number, 'Tipo:', sar.get_report_type_display()],
            ['Fecha:', sar.detected_at.strftime('%d/%m/%Y %H:%M'), 'Estado:', sar.get_status_display()],
            ['Detectado por:', sar.detected_by.get_full_name() or sar.detected_by.username,
             'Monto:', f'{sar.currency_involved} {float(sar.amount_involved):,.2f}'],
        ]
        t_meta = Table(meta, colWidths=[3.5*cm,5.5*cm,3.5*cm,5.5*cm])
        t_meta.setStyle(TableStyle([
            ('FONTNAME', (0,0),(0,-1), 'Helvetica-Bold'),
            ('FONTNAME', (2,0),(2,-1), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0),(-1,-1), 9),
            ('BACKGROUND', (0,0),(0,-1), ASFI_LIGHT),
            ('BACKGROUND', (2,0),(2,-1), ASFI_LIGHT),
            ('GRID', (0,0),(-1,-1), 0.5, colors.HexColor('#CCCCCC')),
            ('BOTTOMPADDING', (0,0),(-1,-1), 6),
            ('TOPPADDING', (0,0),(-1,-1), 6),
        ]))
        story.append(t_meta)
        story.append(Spacer(1, 0.4*cm))

        c = sar.customer
        story.append(Paragraph('DATOS DEL CLIENTE',
            ParagraphStyle('H2', parent=styles['Heading2'], textColor=ASFI_DARK, fontSize=11)))
        t_c = Table([
            ['Nombre:',    c.full_name],
            ['Documento:', f'{c.document_type}: {c.document_number}'],
            ['Telefono:',  c.phone or '-'],
            ['PEP:',       'SI - PERSONA EXPUESTA POLITICAMENTE' if c.is_pep else 'No'],
        ], colWidths=[3.5*cm,14.5*cm])
        t_c.setStyle(TableStyle([
            ('FONTNAME', (0,0),(0,-1), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0),(-1,-1), 9),
            ('BACKGROUND', (0,0),(0,-1), ASFI_LIGHT),
            ('GRID', (0,0),(-1,-1), 0.5, colors.HexColor('#CCCCCC')),
            ('BOTTOMPADDING', (0,0),(-1,-1), 5),
            ('TOPPADDING', (0,0),(-1,-1), 5),
        ]))
        story.append(t_c)
        story.append(Spacer(1, 0.4*cm))

        story.append(Paragraph('DESCRIPCION DE LA ACTIVIDAD',
            ParagraphStyle('H2', parent=styles['Heading2'], textColor=ASFI_DARK, fontSize=11)))
        story.append(Paragraph(sar.description,
            ParagraphStyle('B', parent=styles['Normal'], fontSize=9, leading=14)))

        story.append(Spacer(1, 1*cm))
        story.append(HRFlowable(width='100%', thickness=1, color=ASFI_DARK))
        story.append(Paragraph('DOCUMENTO ESTRICTAMENTE CONFIDENCIAL - USO EXCLUSIVO ASFI',
            ParagraphStyle('CONF', parent=styles['Normal'], fontSize=10,
                           fontName='Helvetica-Bold', textColor=RED_CONF, alignment=TA_CENTER)))
        doc.build(story)

        if user:
            from reports.models import GeneratedReport
            GeneratedReport.objects.create(
                report_type='ROUE_REPORT', format='PDF',
                date_from=sar.detected_at.date(), date_to=sar.detected_at.date(),
                file_path=path.replace(settings.MEDIA_ROOT,'').lstrip('/'),
                file_size_kb=os.path.getsize(path)//1024,
                generated_by=user, parameters={'sar_id': sar_id},
            )
        return path

    # ── PEP ───────────────────────────────────────────────────────────────────

    @classmethod
    def generate_pep_report(cls, year: int = None, month: int = None, user=None) -> dict:
        from reports.models import PEPRegistry, GeneratedReport
        import calendar

        today = date.today()
        if year is None:
            year = today.year
        if month is None:
            month = today.month

        report_date = date(year, month, 1)
        month_end   = date(year, month, calendar.monthrange(year, month)[1])

        # Filter PEPs active during the requested month:
        # since_date <= last_day_of_month AND (until_date IS NULL OR until_date >= first_day_of_month)
        peps = (PEPRegistry.objects
                .select_related('customer')
                .filter(since_date__lte=month_end)
                .filter(models.Q(until_date__isnull=True) | models.Q(until_date__gte=report_date))
                .order_by('risk_level', 'customer__full_name'))

        total_peps = peps.count()
        audit_log.info(
            'ASFI_PEP_GENERATE year=%d month=%d total_peps=%d user=%s',
            year, month, total_peps,
            getattr(user, 'username', 'system'),
        )

        excel_path = cls._pep_excel(peps, report_date, year, month)
        pdf_path   = cls._pep_pdf(peps, report_date, year, month)
        if user:
            for path, fmt in [(excel_path,'EXCEL'),(pdf_path,'PDF')]:
                GeneratedReport.objects.create(
                    report_type='PEP_LIST', format=fmt,
                    date_from=report_date, date_to=month_end,
                    file_path=path.replace(settings.MEDIA_ROOT,'').lstrip('/'),
                    file_size_kb=os.path.getsize(path)//1024,
                    generated_by=user,
                    parameters={'year': year, 'month': month, 'total_peps': total_peps},
                )
        return {
            'status':      'ok',
            'year':        year,
            'month':       month,
            'total_peps':  total_peps,
            'excel_path':  excel_path,
            'pdf_path':    pdf_path,
        }

    @classmethod
    def _pep_excel(cls, peps, report_date, year, month) -> str:
        import calendar
        month_end = date(year, month, calendar.monthrange(year, month)[1])
        period_label = report_date.strftime('%B %Y').upper()

        wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Registro PEP'
        thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                      top=Side(style='thin'), bottom=Side(style='thin'))
        ca = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for merge, text, sz in [
            ('A1:J1', 'AUTORIDAD DE SUPERVISION DEL SISTEMA FINANCIERO - ASFI', 13),
            ('A2:J2', f'REGISTRO DE PERSONAS EXPUESTAS POLITICAMENTE (PEP) - {period_label}', 11),
            ('A3:J3', f'Periodo: {report_date.strftime("%d/%m/%Y")} al {month_end.strftime("%d/%m/%Y")}', 10),
        ]:
            ws.merge_cells(merge)
            cell = ws[merge.split(':')[0]]
            cell.value = text
            cell.font  = Font(bold=True, size=sz, color='003366')
            cell.alignment = ca

        _xl_header(ws, 5, ['Cliente','Documento','Cargo','Institucion',
                             'Desde','Hasta','Riesgo','DD Reforzada','Prox. Revision','Estado'])
        for row_i, pep in enumerate(peps, 6):
            rc = {'HIGH':'FFE0E0','MEDIUM':'FFFDE0','LOW':'E0FFE0'}.get(pep.risk_level,'FFFFFF')
            vals = [
                pep.customer.full_name,
                f'{pep.customer.document_type}: {pep.customer.document_number}',
                pep.position, pep.institution,
                pep.since_date.strftime('%d/%m/%Y'),
                pep.until_date.strftime('%d/%m/%Y') if pep.until_date else 'Activo',
                pep.risk_level, 'SI' if pep.enhanced_dd else 'NO',
                pep.review_date.strftime('%d/%m/%Y'),
                'ACTIVO' if pep.is_active else 'INACTIVO',
            ]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row_i, column=col, value=val)
                cell.border = thin; cell.alignment = ca
                cell.fill   = PatternFill('solid', fgColor=rc)

        tr = peps.count() + 6
        ws.merge_cells(f'A{tr}:I{tr}')
        ws[f'A{tr}']      = f'TOTAL: {peps.count()} registros PEP activos en el periodo'
        ws[f'A{tr}'].font = Font(bold=True, color='003366')

        for i, w in enumerate([28,22,25,25,12,12,10,14,14,10], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        path = os.path.join(_ensure(REPORTS_DIR), f'PEP_{year}{month:02d}.xlsx')
        wb.save(path); return path

    @classmethod
    def _pep_pdf(cls, peps, report_date, year, month) -> str:
        import calendar
        month_end    = date(year, month, calendar.monthrange(year, month)[1])
        period_label = report_date.strftime('%B %Y').upper()

        path   = os.path.join(_ensure(REPORTS_DIR), f'PEP_{year}{month:02d}.pdf')
        doc    = SimpleDocTemplate(path, pagesize=A4, topMargin=1.5*cm, bottomMargin=1.5*cm,
                                    leftMargin=1.5*cm, rightMargin=1.5*cm)
        styles = getSampleStyleSheet(); story = []
        story.append(Paragraph('AUTORIDAD DE SUPERVISION DEL SISTEMA FINANCIERO',
            ParagraphStyle('T1', parent=styles['Title'], textColor=ASFI_DARK, fontSize=13)))
        story.append(Paragraph(
            f'REGISTRO DE PERSONAS EXPUESTAS POLITICAMENTE (PEP) - {period_label}',
            ParagraphStyle('T2', parent=styles['Normal'], textColor=ASFI_DARK,
                           fontSize=10, alignment=TA_CENTER, spaceAfter=2)))
        story.append(Paragraph(
            f'Periodo: {report_date.strftime("%d/%m/%Y")} al {month_end.strftime("%d/%m/%Y")} '
            f'| Total: {peps.count()} registros',
            ParagraphStyle('S', parent=styles['Normal'], fontSize=9,
                           textColor=colors.grey, alignment=TA_CENTER, spaceAfter=8)))
        story.append(HRFlowable(width='100%', thickness=1.5, color=ASFI_DARK))
        story.append(Spacer(1, 0.3*cm))
        data = [['Cliente','Documento','Cargo / Institucion','Desde','Riesgo','DD Ref.']]
        for pep in peps:
            data.append([
                pep.customer.full_name[:28],
                f'{pep.customer.document_type}\n{pep.customer.document_number}',
                f'{pep.position[:28]}\n{pep.institution[:28]}',
                pep.since_date.strftime('%d/%m/%Y'),
                pep.risk_level, 'SI' if pep.enhanced_dd else 'NO',
            ])
        t = Table(data, colWidths=[4.5*cm,3*cm,6*cm,2.2*cm,1.8*cm,1.5*cm], repeatRows=1)
        t.setStyle(_pdf_table_style())
        story.append(t)
        story.append(Spacer(1, 0.5*cm))
        story.append(Paragraph(
            f'Generado el {datetime.now().strftime("%d/%m/%Y %H:%M")} - DOCUMENTO CONFIDENCIAL ASFI',
            ParagraphStyle('F', parent=styles['Normal'],
                           fontSize=7, textColor=colors.grey, alignment=TA_CENTER)))
        doc.build(story); return path

    # ── Libro Diario ──────────────────────────────────────────────────────────

    @classmethod
    def generate_daily_log(cls, log_date: date, branch_id: int, user=None) -> dict:
        from reports.models import DailyOperationLog, CashTransactionReport, GeneratedReport
        from transactions.models import Transaction
        from django.utils import timezone

        log, _ = DailyOperationLog.objects.get_or_create(
            log_date=log_date, branch_id=branch_id)

        # ── Conteo total completadas (sin filtro ASFI) para calcular excluidas ─
        base_qs = Transaction.objects.filter(
            created_at__date=log_date,
            branch_id=branch_id,
            status='COMPLETED',
        )
        total_universe = base_qs.count()

        # ── Queryset filtrado: SOLO transacciones con visible_asfi=True ────────
        txs = (base_qs
               .filter(visible_asfi=True)
               .select_related('customer', 'currency_from', 'currency_to', 'cashier'))

        total_included = txs.count()
        total_excluded = total_universe - total_included

        # ── Validación pre-generación: ninguna transacción interna puede ──────
        # haber filtrado al queryset resultante.
        leaked = txs.filter(visible_asfi=False).count()
        if leaked > 0:
            audit_log.critical(
                'ASFI_DAILY_LOG_VALIDATION_FAIL date=%s branch_id=%d leaked_internal=%d '
                'user=%s — reporte NO generado',
                log_date, branch_id, leaked,
                getattr(user, 'username', 'system'),
            )
            raise ValueError(
                f'Validación ASFI fallida: {leaked} transacción(es) interna(s) '
                f'detectada(s) en el queryset Libro Diario {log_date} sucursal {branch_id}. '
                f'El reporte no fue generado.'
            )

        # ── Log de auditoría ──────────────────────────────────────────────────
        audit_log.info(
            'ASFI_DAILY_LOG_GENERATE date=%s branch_id=%d included=%d '
            'excluded_internal=%d user=%s',
            log_date, branch_id, total_included, total_excluded,
            getattr(user, 'username', 'system'),
        )

        agg = txs.aggregate(
            count     =Count('id'),
            total_buy =Sum('amount_to', filter=Q(transaction_type='BUY')),
            total_sell=Sum('amount_to', filter=Q(transaction_type='SELL')),
        )
        log.total_transactions = agg['count']     or 0
        log.total_buy_bob      = agg['total_buy']  or Decimal('0')
        log.total_sell_bob     = agg['total_sell'] or Decimal('0')
        log.total_profit_bob   = log.total_sell_bob - log.total_buy_bob
        log.rte_count = CashTransactionReport.objects.filter(report_date=log_date).count()

        excel_path = cls._daily_excel(log, txs, log_date)
        pdf_path   = cls._daily_pdf(log, txs, log_date)

        log.excel_file = excel_path; log.pdf_file = pdf_path; log.status = 'CLOSED'
        if user:
            log.closed_by = user; log.closed_at = timezone.now()
        log.save()

        if user:
            for path, fmt in [(excel_path,'EXCEL'),(pdf_path,'PDF')]:
                GeneratedReport.objects.create(
                    report_type='DAILY_LOG', format=fmt,
                    date_from=log_date, date_to=log_date,
                    file_path=path.replace(settings.MEDIA_ROOT,'').lstrip('/'),
                    file_size_kb=os.path.getsize(path)//1024,
                    generated_by=user,
                    parameters={
                        'branch_id':      branch_id,
                        'total_included': total_included,
                        'total_excluded': total_excluded,
                    },
                )
        return {
            'status':             'ok',
            'date':               str(log_date),
            'total_transactions': log.total_transactions,
            'total_excluded':     total_excluded,
            'total_profit_bob':   float(log.total_profit_bob),
            'rte_count':          log.rte_count,
            'excel_path':         excel_path,
            'pdf_path':           pdf_path,
        }

    @classmethod
    def _daily_excel(cls, log, txs, log_date) -> str:
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = f'Libro {log_date}'
        thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                      top=Side(style='thin'), bottom=Side(style='thin'))
        ca = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:L1')
        ws['A1'] = f'LIBRO DIARIO DE OPERACIONES - {log_date.strftime("%d/%m/%Y")}'
        ws['A1'].font = Font(bold=True, size=13, color='003366'); ws['A1'].alignment = ca
        for col,(label,val) in enumerate([
            ('Total trans.',  log.total_transactions),
            ('Compras BOB',   f'Bs. {float(log.total_buy_bob):,.2f}'),
            ('Ventas BOB',    f'Bs. {float(log.total_sell_bob):,.2f}'),
            ('Utilidad',      f'Bs. {float(log.total_profit_bob):,.2f}'),
            ('Registros RTE', log.rte_count),
        ], 1):
            ws.cell(row=3,column=col,value=label).font = Font(bold=True,color='003366')
            ws.cell(row=4,column=col,value=str(val)).font = Font(bold=True,size=11)
        _xl_header(ws, 6, ['N Trans.','Hora','Tipo','Cliente','Documento',
                             'Divisa','Monto','Tasa','Total BOB','Pago','Cajero','RTE'])
        for row_i, tx in enumerate(txs, 7):
            from reports.models import CashTransactionReport as RTEModel
            is_rte = RTEModel.objects.filter(transaction=tx).exists()
            vals = [
                tx.transaction_number, tx.created_at.strftime('%H:%M'),
                tx.get_transaction_type_display(), tx.customer.full_name[:20],
                tx.customer.document_number, str(tx.currency_from),
                f'{float(tx.amount_from):,.2f}', f'{float(tx.exchange_rate):,.4f}',
                f'{float(tx.amount_to):,.2f}', tx.get_payment_method_display(),
                tx.cashier.get_full_name() or tx.cashier.username,
                'RTE' if is_rte else '',
            ]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row_i, column=col, value=val)
                cell.border = thin; cell.alignment = ca
            if is_rte:
                ws.cell(row=row_i,column=12).fill = PatternFill('solid',fgColor='FFE0B2')
        for i,w in enumerate([18,8,8,20,16,10,12,10,14,14,16,8],1):
            ws.column_dimensions[get_column_letter(i)].width = w
        path = os.path.join(_ensure(REPORTS_DIR),
                            f'LIBRO_DIARIO_{log_date.strftime("%Y%m%d")}.xlsx')
        wb.save(path); return path

    @classmethod
    def _daily_pdf(cls, log, txs, log_date) -> str:
        path   = os.path.join(_ensure(REPORTS_DIR),
                              f'LIBRO_DIARIO_{log_date.strftime("%Y%m%d")}.pdf')
        doc    = SimpleDocTemplate(path, pagesize=A4, topMargin=1.5*cm, bottomMargin=1.5*cm,
                                    leftMargin=1*cm, rightMargin=1*cm)
        styles = getSampleStyleSheet(); story = []
        h2     = ParagraphStyle('H2', parent=styles['Heading2'], textColor=ASFI_DARK, fontSize=11)
        story.append(Paragraph(f'LIBRO DIARIO DE OPERACIONES - {log_date.strftime("%d/%m/%Y")}',
            ParagraphStyle('T',parent=styles['Title'],textColor=ASFI_DARK,fontSize=14)))
        story.append(HRFlowable(width='100%',thickness=2,color=ASFI_DARK))
        story.append(Spacer(1,0.3*cm))
        t_sum = Table([
            ['Total trans.', str(log.total_transactions),
             'Compras (BOB)', f'Bs. {float(log.total_buy_bob):,.2f}'],
            ['Ventas (BOB)', f'Bs. {float(log.total_sell_bob):,.2f}',
             'Utilidad', f'Bs. {float(log.total_profit_bob):,.2f}'],
        ], colWidths=[3.5*cm,5*cm,3.5*cm,6*cm])
        t_sum.setStyle(TableStyle([
            ('FONTNAME',(0,0),(0,-1),'Helvetica-Bold'),
            ('FONTNAME',(2,0),(2,-1),'Helvetica-Bold'),
            ('BACKGROUND',(0,0),(0,-1),ASFI_LIGHT),
            ('BACKGROUND',(2,0),(2,-1),ASFI_LIGHT),
            ('GRID',(0,0),(-1,-1),0.5,colors.HexColor('#CCCCCC')),
            ('FONTSIZE',(0,0),(-1,-1),9),
            ('BOTTOMPADDING',(0,0),(-1,-1),6),
        ]))
        story.append(t_sum); story.append(Spacer(1,0.4*cm))
        story.append(Paragraph('DETALLE DE OPERACIONES',h2))
        tx_data = [['N Trans.','Hora','Tipo','Cliente','Monto','Tasa','BOB','Pago']]
        for tx in txs:
            tx_data.append([
                tx.transaction_number, tx.created_at.strftime('%H:%M'),
                tx.get_transaction_type_display(), tx.customer.full_name[:20],
                f'{float(tx.amount_from):,.2f}', f'{float(tx.exchange_rate):,.4f}',
                f'{float(tx.amount_to):,.2f}', tx.get_payment_method_display(),
            ])
        t = Table(tx_data,
                  colWidths=[2.8*cm,1.5*cm,1.5*cm,4.5*cm,2.5*cm,2*cm,2.5*cm,2.2*cm],
                  repeatRows=1)
        t.setStyle(_pdf_table_style())
        story.append(t); doc.build(story); return path
    # ── Export CSV (RTE mensual y Libro Diario) ───────────────────────────────

    @classmethod
    def generate_rte_csv(cls, year: int, month: int, user=None) -> str:
        """
        CSV del RTE mensual — mismo filtro visible_asfi y validación que
        Excel/PDF. Retorna la ruta del archivo generado (UTF-8 con BOM para
        que Excel abra bien los acentos).
        """
        from reports.models import CashTransactionReport, GeneratedReport
        import calendar
        import csv

        start = date(year, month, 1)
        end   = date(year, month, calendar.monthrange(year, month)[1])

        rtes = (CashTransactionReport.objects
                .filter(report_date__gte=start, report_date__lte=end,
                        transaction__visible_asfi=True)
                .select_related('transaction', 'transaction__customer')
                .order_by('report_date'))

        leaked = rtes.filter(transaction__visible_asfi=False).count()
        if leaked > 0:
            audit_log.critical(
                'ASFI_RTE_CSV_VALIDATION_FAIL year=%d month=%d leaked_internal=%d '
                'user=%s — reporte NO generado',
                year, month, leaked, getattr(user, 'username', 'system'),
            )
            raise ValueError(
                f'Validación ASFI fallida: {leaked} transacción(es) interna(s) '
                f'detectada(s) en el queryset RTE {year}-{month:02d}. '
                f'El reporte no fue generado.'
            )

        audit_log.info(
            'ASFI_RTE_CSV_GENERATE year=%d month=%d included=%d user=%s',
            year, month, rtes.count(), getattr(user, 'username', 'system'),
        )

        path = os.path.join(_ensure(REPORTS_DIR), f'RTE_{year}{month:02d}.csv')
        total_usd = Decimal('0')
        with open(path, 'w', encoding='utf-8-sig', newline='') as f:
            w = csv.writer(f)
            w.writerow(['N RTE', 'Fecha', 'N Transaccion', 'Tipo',
                        'Cliente', 'Tipo Doc.', 'N Documento', 'Nacionalidad',
                        'Monto Original', 'Divisa', 'Equiv. USD', 'PEP'])
            for rte in rtes:
                tx = rte.transaction
                w.writerow([
                    rte.report_number,
                    rte.report_date.strftime('%d/%m/%Y'),
                    tx.transaction_number,
                    'COMPRA' if tx.transaction_type == 'BUY' else 'VENTA',
                    rte.customer_full_name,
                    rte.customer_document_type,
                    rte.customer_document_num,
                    rte.customer_nationality,
                    f'{rte.original_amount:.2f}',
                    rte.currency_code,
                    f'{rte.amount_usd_equiv:.2f}',
                    'SI' if rte.customer_is_pep else 'NO',
                ])
                total_usd += rte.amount_usd_equiv
            w.writerow([])
            w.writerow(['TOTAL USD', f'{total_usd:.2f}'])

        if user:
            GeneratedReport.objects.create(
                report_type='RTE_MONTHLY', format='CSV',
                date_from=start, date_to=end,
                file_path=path.replace(settings.MEDIA_ROOT, '').lstrip('/'),
                file_size_kb=os.path.getsize(path) // 1024,
                generated_by=user,
                parameters={'year': year, 'month': month},
            )
        return path

    @classmethod
    def generate_daily_log_csv(cls, log_date: date, branch_id: int, user=None) -> str:
        """
        CSV del Libro Diario — mismo filtro visible_asfi que Excel/PDF.
        Retorna la ruta del archivo generado.
        """
        from reports.models import CashTransactionReport, GeneratedReport
        from transactions.models import Transaction
        import csv

        txs = (Transaction.objects
               .filter(created_at__date=log_date, branch_id=branch_id,
                       status='COMPLETED', visible_asfi=True)
               .select_related('customer', 'currency_from', 'currency_to',
                               'cashier')
               .order_by('created_at'))

        audit_log.info(
            'ASFI_DAILY_LOG_CSV_GENERATE date=%s branch_id=%d included=%d user=%s',
            log_date, branch_id, txs.count(), getattr(user, 'username', 'system'),
        )

        rte_tx_ids = set(
            CashTransactionReport.objects
            .filter(report_date=log_date)
            .values_list('transaction_id', flat=True)
        )

        path = os.path.join(
            _ensure(REPORTS_DIR),
            f'LIBRO_DIARIO_{log_date.strftime("%Y%m%d")}.csv')
        with open(path, 'w', encoding='utf-8-sig', newline='') as f:
            w = csv.writer(f)
            w.writerow(['N Trans.', 'Hora', 'Tipo', 'Cliente', 'Documento',
                        'Divisa', 'Monto', 'Tasa', 'Total BOB', 'Pago',
                        'Cajero', 'RTE'])
            for tx in txs:
                w.writerow([
                    tx.transaction_number,
                    tx.created_at.strftime('%H:%M'),
                    tx.get_transaction_type_display(),
                    tx.customer.full_name if tx.customer_id else '',
                    tx.customer.document_number if tx.customer_id else '',
                    str(tx.currency_from),
                    f'{float(tx.amount_from):.2f}',
                    f'{float(tx.exchange_rate):.4f}',
                    f'{float(tx.amount_to):.2f}',
                    tx.get_payment_method_display(),
                    tx.cashier.get_full_name() or tx.cashier.username,
                    'RTE' if tx.pk in rte_tx_ids else '',
                ])

        if user:
            GeneratedReport.objects.create(
                report_type='DAILY_LOG', format='CSV',
                date_from=log_date, date_to=log_date,
                file_path=path.replace(settings.MEDIA_ROOT, '').lstrip('/'),
                file_size_kb=os.path.getsize(path) // 1024,
                generated_by=user,
                parameters={'branch_id': branch_id},
            )
        return path
